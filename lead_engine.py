"""Lead Engine v2 — local lead collection app for a commercial loan brokerage.

One Flask application in one file.  It collects leads from free public
data sources.  It removes duplicates.  It does checks on each lead.  It
gives each lead a score from 0 to 100.  It exports CSV files.

Run:  python lead_engine.py [--demo] [--port 5002]
The launcher (launcher.pyw) starts this file with no console window.
"""

import argparse
import csv
import hashlib
import io
import json
import os
import re
import socket
import sqlite3
import sys
import tempfile
import threading
import time
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
from flask import (
    Flask,
    Response,
    jsonify,
    redirect,
    render_template_string,
    request,
    url_for,
)
from jinja2 import DictLoader

from contact_scraper import USER_AGENT, scrape_firm

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR / "lead_engine.db"
OLD_DB_PATH = APP_DIR / "leads.db"
SERVICE_ACCOUNT_PATH = APP_DIR / "service_account.json"
IMPORT_DIR = APP_DIR / "import_uploads"

HOST = "127.0.0.1"
DEFAULT_PORT = 5002

SUPPORTED_STATES = ["WI", "ND"]
STATE_NAMES = {"WI": "Wisconsin", "ND": "North Dakota"}

REQUEST_TIMEOUT = 30
DOWNLOAD_TIMEOUT = 600
POLITE_DELAY = 2.0

# Source pages.  Each connector fails with a readable message if a page
# or a format changes.
SEC_ADV_PAGE = (
    "https://www.sec.gov/foia-services/frequently-requested-documents/"
    "form-adv-data"
)
FDIC_API = "https://banks.data.fdic.gov/api/institutions"
NCUA_QUARTERLY_PAGE = (
    "https://ncua.gov/analysis/credit-union-corporate-call-report-data/"
    "quarterly-data"
)
IRS_PTIN_PAGE = (
    "https://www.irs.gov/tax-professionals/"
    "ptin-information-and-the-freedom-of-information-act"
)
OVERPASS_API = "https://overpass-api.de/api/interpreter"
SBA_CKAN_API = "https://data.sba.gov/api/3/action/package_show?id=7-a-504-foia"

# Connectors whose rows come from an official registry.
REGISTRY_CONNECTORS = {
    "ria_sec", "ria_state", "fdic_banks", "ncua_cus", "irs_ptin",
    "sba_borrowers", "migrated",
}

# Domains that many firms share.  Do not use them for duplicate matches.
SHARED_DOMAINS = {
    "facebook.com", "linkedin.com", "google.com", "yelp.com",
    "instagram.com", "twitter.com", "x.com", "youtube.com",
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "aol.com",
    "wixsite.com", "squarespace.com", "godaddysites.com",
}

FREE_EMAIL_DOMAINS = {
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "aol.com",
    "icloud.com", "msn.com", "sbcglobal.net", "att.net", "comcast.net",
    "charter.net", "live.com", "protonmail.com", "mail.com", "me.com",
}

DEFAULT_WEIGHTS = {
    "weight_registry": 30,
    "weight_website": 20,
    "weight_phone": 15,
    "weight_email": 15,
    "weight_address": 10,
    "weight_cross": 10,
}
WEIGHT_LABELS = {
    "weight_registry": "Registry source",
    "weight_website": "Live website",
    "weight_phone": "Valid phone",
    "weight_email": "Email domain has MX",
    "weight_address": "Address complete",
    "weight_cross": "Cross-source match",
}
CHECK_TO_WEIGHT = {
    "registry": "weight_registry",
    "website": "weight_website",
    "phone": "weight_phone",
    "email_mx": "weight_email",
    "address": "weight_address",
    "cross_source": "weight_cross",
}

LEAD_TYPES = {
    "ria_sec": "Financial advisor (SEC RIA)",
    "ria_state": "Financial advisor (state)",
    "bank": "Community bank",
    "credit_union": "Credit union",
    "cpa": "CPA / tax preparer",
    "insurance": "Insurance agency",
    "cre_agent": "Commercial RE agent",
    "attorney": "Business attorney",
    "borrower": "Borrower prospect",
    "generic": "Generic import",
}

IMPORT_FIELDS = [
    ("firm_name", "Firm name"),
    ("dba_name", "DBA name"),
    ("contact_name", "Contact name"),
    ("address", "Street address"),
    ("city", "City"),
    ("state", "State"),
    ("zip_code", "ZIP code"),
    ("phone", "Phone"),
    ("email", "Email"),
    ("website", "Website"),
    ("external_id", "External ID / license #"),
    ("aum_dollars", "AUM ($)"),
    ("num_employees", "Employees"),
    ("source_detail", "Source detail"),
    ("notes", "Notes"),
]

COMPLIANCE_NOTE = (
    "Compliance: exported lists that you use for email or phone outreach "
    "are subject to CAN-SPAM and the TCPA. Referral-fee agreements with "
    "CPAs, banks, and advisors need attorney review before use."
)


# ---------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------

def get_db():
    """Open a database connection for the current call."""
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


SCHEMA = """
CREATE TABLE IF NOT EXISTS leads (
    id INTEGER PRIMARY KEY,
    lead_type TEXT NOT NULL DEFAULT 'generic',
    external_id TEXT DEFAULT '',
    firm_name TEXT DEFAULT '',
    dba_name TEXT DEFAULT '',
    contact_name TEXT DEFAULT '',
    address TEXT DEFAULT '',
    city TEXT DEFAULT '',
    state TEXT DEFAULT '',
    zip_code TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    phone_normalized TEXT DEFAULT '',
    email TEXT DEFAULT '',
    email_domain TEXT DEFAULT '',
    website TEXT DEFAULT '',
    website_domain TEXT DEFAULT '',
    aum_dollars INTEGER,
    num_employees INTEGER,
    source_connector TEXT DEFAULT '',
    source_detail TEXT DEFAULT '',
    corroborators TEXT DEFAULT '',
    score INTEGER DEFAULT 0,
    tier TEXT DEFAULT 'C',
    first_seen TEXT DEFAULT '',
    last_seen TEXT DEFAULT '',
    status TEXT DEFAULT 'new',
    notes TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS contacts (
    id INTEGER PRIMARY KEY,
    lead_id INTEGER NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
    name TEXT DEFAULT '',
    title TEXT DEFAULT '',
    email TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    linkedin_url TEXT DEFAULT '',
    source_url TEXT DEFAULT '',
    confidence REAL DEFAULT 0,
    scraped_at TEXT DEFAULT '',
    notes TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS verifications (
    id INTEGER PRIMARY KEY,
    lead_id INTEGER NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
    check_name TEXT NOT NULL,
    passed INTEGER NOT NULL DEFAULT 0,
    detail TEXT DEFAULT '',
    checked_at TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS dedup_queue (
    id INTEGER PRIMARY KEY,
    lead_id_a INTEGER NOT NULL,
    lead_id_b INTEGER NOT NULL,
    match_reason TEXT DEFAULT '',
    similarity REAL DEFAULT 0,
    resolved INTEGER DEFAULT 0,
    resolution TEXT DEFAULT '',
    created_at TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS runs (
    id INTEGER PRIMARY KEY,
    connector_id TEXT NOT NULL,
    started_at TEXT DEFAULT '',
    finished_at TEXT DEFAULT '',
    rows_seen INTEGER DEFAULT 0,
    rows_added INTEGER DEFAULT 0,
    rows_updated INTEGER DEFAULT 0,
    rows_deduped INTEGER DEFAULT 0,
    error TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS lenders (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    city TEXT DEFAULT '',
    state TEXT DEFAULT '',
    loan_count INTEGER DEFAULT 0,
    total_dollars INTEGER DEFAULT 0,
    last_seen TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS scrape_log (
    id INTEGER PRIMARY KEY,
    lead_id INTEGER NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
    scraped_at TEXT DEFAULT '',
    ok INTEGER DEFAULT 0,
    contacts_found INTEGER DEFAULT 0,
    error TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_leads_external ON leads(external_id);
CREATE INDEX IF NOT EXISTS idx_leads_phone ON leads(phone_normalized);
CREATE INDEX IF NOT EXISTS idx_leads_domain ON leads(website_domain);
CREATE INDEX IF NOT EXISTS idx_leads_name_city ON leads(firm_name, city);
CREATE INDEX IF NOT EXISTS idx_contacts_lead ON contacts(lead_id);
CREATE INDEX IF NOT EXISTS idx_verif_lead ON verifications(lead_id);
CREATE UNIQUE INDEX IF NOT EXISTS idx_lenders_name ON lenders(name, state);
CREATE VIEW IF NOT EXISTS all_leads AS
    SELECT l.*,
        (SELECT COUNT(*) FROM contacts c WHERE c.lead_id = l.id)
            AS contact_count,
        (SELECT COUNT(*) FROM verifications v
            WHERE v.lead_id = l.id AND v.passed = 1) AS checks_passed,
        (SELECT COUNT(*) FROM verifications v
            WHERE v.lead_id = l.id) AS checks_total
    FROM leads l;
"""


def init_db():
    """Create all tables, indexes, and the view.  Safe to run twice."""
    conn = get_db()
    try:
        conn.executescript(SCHEMA)
        conn.commit()
    finally:
        conn.close()


def now_iso():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_setting(conn, key, default=""):
    row = conn.execute(
        "SELECT value FROM settings WHERE key = ?", (key,)
    ).fetchone()
    return row["value"] if row else default


def set_setting(conn, key, value):
    conn.execute(
        "INSERT INTO settings(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, str(value)),
    )


def get_weights(conn):
    weights = {}
    for key, default in DEFAULT_WEIGHTS.items():
        raw = get_setting(conn, key, str(default))
        try:
            weights[key] = max(0, int(float(raw)))
        except ValueError:
            weights[key] = default
    return weights


def get_target_state(conn):
    state = get_setting(conn, "target_state", "WI").upper()
    return state if state in SUPPORTED_STATES else "WI"


# ---------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------

LEGAL_SUFFIXES = {
    "llc", "inc", "incorporated", "corp", "corporation", "ltd", "llp",
    "lp", "pc", "sc", "co", "company", "pllc", "pa", "plc", "ltd.",
    "l.l.c", "l.l.p", "s.c", "p.c",
}

_PUNCT_RE = re.compile(r"[^\w\s]")
_SPACE_RE = re.compile(r"\s+")


def normalize_name(name):
    """Make a firm name comparable.  Remove case, punctuation, and
    legal suffixes such as LLC and Inc."""
    if not name:
        return ""
    text = name.lower().replace("&", " and ")
    text = _PUNCT_RE.sub(" ", text)
    words = _SPACE_RE.sub(" ", text).strip().split(" ")
    while words and words[-1] in LEGAL_SUFFIXES:
        words.pop()
    return " ".join(words)


def normalize_phone(phone):
    """Return ten digits for a valid NANP number, or an empty string."""
    if not phone:
        return ""
    digits = re.sub(r"\D", "", str(phone))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return ""
    if digits[0] in "01" or digits[3] in "01":
        return ""
    return digits


def extract_domain(url_or_email):
    """Return the bare domain from a URL or an email address."""
    if not url_or_email:
        return ""
    text = str(url_or_email).strip().lower()
    if "@" in text and "/" not in text:
        domain = text.rsplit("@", 1)[1]
    else:
        if not text.startswith(("http://", "https://")):
            text = "http://" + text
        try:
            domain = urlparse(text).netloc
        except ValueError:
            return ""
    domain = domain.split(":")[0].strip(".")
    if domain.startswith("www."):
        domain = domain[4:]
    if "." not in domain:
        return ""
    return domain


def to_int(value):
    """Read an integer from messy text.  Return None if there is none."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d\-]", "", str(value).split(".")[0])
    if digits in ("", "-"):
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def first_match(headers, candidates):
    """Find the first header that contains one of the candidate
    substrings.  Candidates are in priority order.  Returns the header
    name or None."""
    lowered = [(h, h.lower()) for h in headers]
    for candidate in candidates:
        cand = candidate.lower()
        for original, low in lowered:
            if cand in low:
                return original
    return None


# ---------------------------------------------------------------------
# Upsert and exact duplicate cascade
# ---------------------------------------------------------------------

UPDATABLE_FIELDS = [
    "firm_name", "dba_name", "contact_name", "address", "city", "state",
    "zip_code", "phone", "phone_normalized", "email", "email_domain",
    "website", "website_domain", "aum_dollars", "num_employees",
    "source_detail",
]


def prepare_row(row):
    """Fill the computed fields of a lead row."""
    out = dict(row)
    out["phone_normalized"] = normalize_phone(out.get("phone", ""))
    out["email_domain"] = extract_domain(out.get("email", ""))
    out["website_domain"] = extract_domain(out.get("website", ""))
    for key in ("firm_name", "city", "state", "zip_code"):
        if out.get(key):
            out[key] = " ".join(str(out[key]).split())
    if out.get("state"):
        out["state"] = out["state"].upper()[:2]
    return out


def find_existing(conn, row):
    """Run the exact-match cascade.  Return (lead_row, reason) or
    (None, None).  The cascade order is: external ID, phone, website
    domain, then name plus city."""
    lead_type = row.get("lead_type", "")
    ext = (row.get("external_id") or "").strip()
    if ext:
        hit = conn.execute(
            "SELECT * FROM leads WHERE external_id = ? AND lead_type = ?",
            (ext, lead_type),
        ).fetchone()
        if hit:
            return hit, "external_id"
    phone = row.get("phone_normalized", "")
    if phone:
        hit = conn.execute(
            "SELECT * FROM leads WHERE phone_normalized = ?", (phone,)
        ).fetchone()
        if hit:
            return hit, "phone"
    domain = row.get("website_domain", "")
    if domain and domain not in SHARED_DOMAINS:
        hit = conn.execute(
            "SELECT * FROM leads WHERE website_domain = ?", (domain,)
        ).fetchone()
        if hit:
            return hit, "domain"
    norm = normalize_name(row.get("firm_name", ""))
    city = (row.get("city") or "").strip().lower()
    if norm and city:
        for cand in conn.execute(
            "SELECT * FROM leads WHERE lower(city) = ?", (city,)
        ):
            if normalize_name(cand["firm_name"]) == norm:
                return cand, "name_city"
    return None, None


def upsert_lead(conn, row):
    """Insert a lead, or update the match that the cascade found.

    Returns (lead_id, action).  Action is one of:
      added    — a new row
      updated  — a match by external ID from the same connector
      deduped  — a match through another key or another connector
    An update writes only fields whose new value is not empty.
    """
    row = prepare_row(row)
    existing, reason = find_existing(conn, row)
    stamp = now_iso()
    if existing is None:
        conn.execute(
            """INSERT INTO leads(lead_type, external_id, firm_name,
                dba_name, contact_name, address, city, state, zip_code,
                phone, phone_normalized, email, email_domain, website,
                website_domain, aum_dollars, num_employees,
                source_connector, source_detail, corroborators,
                first_seen, last_seen, status, notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                row.get("lead_type", "generic"),
                row.get("external_id", ""),
                row.get("firm_name", ""),
                row.get("dba_name", ""),
                row.get("contact_name", ""),
                row.get("address", ""),
                row.get("city", ""),
                row.get("state", ""),
                row.get("zip_code", ""),
                row.get("phone", ""),
                row.get("phone_normalized", ""),
                row.get("email", ""),
                row.get("email_domain", ""),
                row.get("website", ""),
                row.get("website_domain", ""),
                row.get("aum_dollars"),
                row.get("num_employees"),
                row.get("source_connector", ""),
                row.get("source_detail", ""),
                "",
                stamp,
                stamp,
                row.get("status") or "new",
                row.get("notes", ""),
            ),
        )
        return conn.execute("SELECT last_insert_rowid() AS i").fetchone()["i"], "added"

    updates = {"last_seen": stamp}
    for field in UPDATABLE_FIELDS:
        new_value = row.get(field)
        if new_value in (None, ""):
            continue
        old_value = existing[field]
        if old_value in (None, ""):
            updates[field] = new_value
        elif field in ("source_detail",) and new_value != old_value:
            updates[field] = new_value
    if not existing["external_id"] and row.get("external_id"):
        updates["external_id"] = row["external_id"]

    action = "updated"
    new_source = row.get("source_connector", "")
    if new_source and new_source != existing["source_connector"]:
        seen = set(
            s for s in (existing["corroborators"] or "").split(",") if s
        )
        if new_source not in seen:
            seen.add(new_source)
            updates["corroborators"] = ",".join(sorted(seen))
        action = "deduped"
    elif reason != "external_id":
        action = "deduped"

    sets = ", ".join(f"{k} = ?" for k in updates)
    conn.execute(
        f"UPDATE leads SET {sets} WHERE id = ?",
        (*updates.values(), existing["id"]),
    )
    return existing["id"], action


# ---------------------------------------------------------------------
# Fuzzy duplicate scan
# ---------------------------------------------------------------------

def run_fuzzy_dedup(conn, threshold=90):
    """Find near-duplicate names in the same city with rapidfuzz.
    Put new pairs in the review queue.  Returns the count of new
    pairs."""
    from rapidfuzz import fuzz

    known = set()
    for row in conn.execute(
        "SELECT lead_id_a, lead_id_b FROM dedup_queue"
    ):
        known.add(frozenset((row["lead_id_a"], row["lead_id_b"])))

    by_city = {}
    for row in conn.execute(
        "SELECT id, firm_name, city FROM leads WHERE firm_name != ''"
    ):
        city = (row["city"] or "").strip().lower()
        if not city:
            continue
        by_city.setdefault(city, []).append(
            (row["id"], normalize_name(row["firm_name"]))
        )

    added = 0
    stamp = now_iso()
    for city, entries in by_city.items():
        n = len(entries)
        for i in range(n):
            id_a, name_a = entries[i]
            if not name_a:
                continue
            for j in range(i + 1, n):
                id_b, name_b = entries[j]
                if not name_b:
                    continue
                pair = frozenset((id_a, id_b))
                if pair in known:
                    continue
                score = fuzz.token_sort_ratio(name_a, name_b)
                if score >= threshold:
                    conn.execute(
                        """INSERT INTO dedup_queue(lead_id_a, lead_id_b,
                            match_reason, similarity, resolved,
                            created_at)
                           VALUES(?,?,?,?,0,?)""",
                        (
                            min(id_a, id_b),
                            max(id_a, id_b),
                            f"Name similarity {score:.0f}% in {city.title()}",
                            round(score, 1),
                            stamp,
                        ),
                    )
                    known.add(pair)
                    added += 1
    conn.commit()
    return added


def count_non_empty(lead_row):
    """Count the fields of a lead that have data."""
    count = 0
    for field in UPDATABLE_FIELDS + ["external_id", "notes"]:
        value = lead_row[field]
        if value not in (None, "", 0):
            count += 1
    return count


def merge_leads(conn, keep_id, drop_id):
    """Merge one lead into another.  The kept lead gets each field that
    it does not have.  Contacts, checks, and log rows move to the kept
    lead.  The other lead is removed."""
    keep = conn.execute("SELECT * FROM leads WHERE id = ?", (keep_id,)).fetchone()
    drop = conn.execute("SELECT * FROM leads WHERE id = ?", (drop_id,)).fetchone()
    if keep is None or drop is None:
        return False
    updates = {}
    for field in UPDATABLE_FIELDS + ["external_id", "notes"]:
        if keep[field] in (None, "") and drop[field] not in (None, ""):
            updates[field] = drop[field]
    sources = set(
        s for s in (keep["corroborators"] or "").split(",") if s
    )
    sources |= set(
        s for s in (drop["corroborators"] or "").split(",") if s
    )
    if drop["source_connector"] and drop["source_connector"] != keep["source_connector"]:
        sources.add(drop["source_connector"])
    updates["corroborators"] = ",".join(sorted(sources))
    if drop["first_seen"] and (
        not keep["first_seen"] or drop["first_seen"] < keep["first_seen"]
    ):
        updates["first_seen"] = drop["first_seen"]
    updates["last_seen"] = now_iso()
    sets = ", ".join(f"{k} = ?" for k in updates)
    conn.execute(
        f"UPDATE leads SET {sets} WHERE id = ?", (*updates.values(), keep_id)
    )
    for table in ("contacts", "verifications", "scrape_log"):
        conn.execute(
            f"UPDATE {table} SET lead_id = ? WHERE lead_id = ?",
            (keep_id, drop_id),
        )
    conn.execute("DELETE FROM leads WHERE id = ?", (drop_id,))
    conn.execute(
        """UPDATE dedup_queue SET resolved = 1, resolution = 'merged'
           WHERE resolved = 0 AND (lead_id_a IN (?, ?) OR lead_id_b IN (?, ?))""",
        (keep_id, drop_id, keep_id, drop_id),
    )
    return True


# ---------------------------------------------------------------------
# Verification and scoring
# ---------------------------------------------------------------------

def check_website(url):
    """Do a HEAD request.  Return (passed, detail)."""
    if not url:
        return 0, "No website on file."
    target = url.strip()
    if not target.startswith(("http://", "https://")):
        target = "https://" + target
    try:
        resp = requests.head(
            target,
            headers={"User-Agent": USER_AGENT},
            timeout=10,
            allow_redirects=True,
        )
        if resp.status_code < 400:
            return 1, f"HTTP {resp.status_code}"
        # Some servers refuse HEAD.  Try a small GET.
        resp = requests.get(
            target,
            headers={"User-Agent": USER_AGENT},
            timeout=10,
            allow_redirects=True,
            stream=True,
        )
        resp.close()
        if resp.status_code < 400:
            return 1, f"HTTP {resp.status_code} (GET)"
        return 0, f"HTTP {resp.status_code}"
    except requests.RequestException as exc:
        return 0, f"No response: {exc.__class__.__name__}"


def check_email_mx(domain):
    """Make sure that the email domain has MX records.  This is a
    domain-level check only, not a mailbox check."""
    if not domain:
        return 0, "No email on file."
    try:
        import dns.resolver

        answers = dns.resolver.resolve(domain, "MX", lifetime=8)
        hosts = ", ".join(
            str(r.exchange).rstrip(".") for r in list(answers)[:2]
        )
        return 1, f"MX: {hosts}"
    except ImportError:
        pass
    except Exception as exc:
        return 0, f"No MX record ({exc.__class__.__name__})"
    # Fallback without dnspython: an A record is a weak signal.
    try:
        socket.gethostbyname(domain)
        return 1, "Domain resolves (A record; dnspython not installed)"
    except OSError:
        return 0, "Domain does not resolve."


def verify_lead(conn, lead, online=True):
    """Run all checks for one lead.  Store one row per check with a
    timestamp.  Then compute the score."""
    stamp = now_iso()
    checks = []

    is_registry = lead["source_connector"] in REGISTRY_CONNECTORS
    corr = set(s for s in (lead["corroborators"] or "").split(",") if s)
    if not is_registry:
        is_registry = bool(corr & REGISTRY_CONNECTORS)
    checks.append((
        "registry",
        1 if is_registry else 0,
        f"Source: {lead['source_connector'] or 'manual import'}",
    ))

    if online:
        passed, detail = check_website(lead["website"])
    else:
        passed, detail = 0, "Not checked yet."
    checks.append(("website", passed, detail))

    phone_ok = 1 if lead["phone_normalized"] else 0
    phone_detail = (
        f"Valid NANP number: {lead['phone_normalized']}"
        if phone_ok
        else "No valid 10-digit phone."
    )
    checks.append(("phone", phone_ok, phone_detail))

    if online:
        passed, detail = check_email_mx(lead["email_domain"])
    else:
        passed, detail = 0, "Not checked yet."
    checks.append(("email_mx", passed, detail))

    parts = [lead["address"], lead["city"], lead["state"], lead["zip_code"]]
    missing = [
        name
        for name, val in zip(("street", "city", "state", "ZIP"), parts)
        if not val
    ]
    checks.append((
        "address",
        0 if missing else 1,
        "Complete address."
        if not missing
        else "Missing: " + ", ".join(missing),
    ))

    checks.append((
        "cross_source",
        1 if corr else 0,
        "Also seen in: " + ", ".join(sorted(corr))
        if corr
        else "One source only.",
    ))

    conn.execute("DELETE FROM verifications WHERE lead_id = ?", (lead["id"],))
    for name, passed, detail in checks:
        conn.execute(
            """INSERT INTO verifications(lead_id, check_name, passed,
                detail, checked_at) VALUES(?,?,?,?,?)""",
            (lead["id"], name, passed, detail, stamp),
        )
    compute_score(conn, lead["id"])


def compute_score(conn, lead_id):
    """Compute the weighted score and the tier for one lead."""
    weights = get_weights(conn)
    total_weight = sum(weights.values()) or 1
    earned = 0
    for row in conn.execute(
        "SELECT check_name, passed FROM verifications WHERE lead_id = ?",
        (lead_id,),
    ):
        weight_key = CHECK_TO_WEIGHT.get(row["check_name"])
        if weight_key and row["passed"]:
            earned += weights[weight_key]
    score = round(100 * earned / total_weight)
    tier = "A" if score >= 75 else ("B" if score >= 50 else "C")
    conn.execute(
        "UPDATE leads SET score = ?, tier = ? WHERE id = ?",
        (score, tier, lead_id),
    )
    return score, tier


def rescore_all(conn):
    """Compute scores again for all leads from stored check rows."""
    ids = [r["id"] for r in conn.execute("SELECT id FROM leads")]
    for lead_id in ids:
        compute_score(conn, lead_id)
    conn.commit()
    return len(ids)


# ---------------------------------------------------------------------
# Migration from the predecessor app
# ---------------------------------------------------------------------

def migrate_predecessor():
    """Copy leads and contacts from an old leads.db file, one time.
    Safe to run twice: a settings flag and the upsert cascade prevent
    duplicate rows."""
    if not OLD_DB_PATH.exists():
        return
    conn = get_db()
    try:
        if get_setting(conn, "migrated_predecessor") == "1":
            return
        old = sqlite3.connect(OLD_DB_PATH)
        old.row_factory = sqlite3.Row
        try:
            tables = {
                r["name"]
                for r in old.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
            crd_to_new = {}
            if "leads" in tables:
                for row in old.execute("SELECT * FROM leads"):
                    data = dict(row)
                    crd = str(
                        data.get("crd") or data.get("external_id") or ""
                    ).strip()
                    new_row = {
                        "lead_type": "ria_sec",
                        "external_id": crd,
                        "source_connector": "migrated",
                        "source_detail": "Migrated from leads.db",
                    }
                    for field in UPDATABLE_FIELDS + ["notes", "status"]:
                        if field in data and data[field] not in (None, ""):
                            new_row[field] = data[field]
                    if not new_row.get("firm_name"):
                        new_row["firm_name"] = str(
                            data.get("name") or data.get("firm") or ""
                        )
                    if not new_row.get("firm_name"):
                        continue
                    lead_id, _ = upsert_lead(conn, new_row)
                    if crd:
                        crd_to_new[crd] = lead_id
            if "contacts" in tables:
                for row in old.execute("SELECT * FROM contacts"):
                    data = dict(row)
                    crd = str(data.get("crd") or data.get("lead_crd") or "").strip()
                    lead_id = crd_to_new.get(crd)
                    if lead_id is None:
                        continue
                    conn.execute(
                        """INSERT INTO contacts(lead_id, name, title,
                            email, phone, linkedin_url, source_url,
                            confidence, scraped_at, notes)
                           VALUES(?,?,?,?,?,?,?,?,?,?)""",
                        (
                            lead_id,
                            str(data.get("name") or ""),
                            str(data.get("title") or ""),
                            str(data.get("email") or ""),
                            str(data.get("phone") or ""),
                            str(
                                data.get("linkedin_url")
                                or data.get("linkedin")
                                or ""
                            ),
                            str(data.get("source_url") or ""),
                            float(data.get("confidence") or 0),
                            str(data.get("scraped_at") or now_iso()),
                            str(data.get("notes") or ""),
                        ),
                    )
        finally:
            old.close()
        set_setting(conn, "migrated_predecessor", "1")
        conn.commit()
    except Exception as exc:
        print(f"Migration from leads.db failed: {exc}", file=sys.stderr)
    finally:
        conn.close()


# ---------------------------------------------------------------------
# Polite HTTP
# ---------------------------------------------------------------------

class ConnectorError(Exception):
    """An error with a message that the UI can show."""


_HOST_LAST_HIT = {}
_HOST_LOCK = threading.Lock()


def _polite_pause(url):
    """Keep at least two seconds between requests to one host."""
    host = urlparse(url).netloc
    with _HOST_LOCK:
        last = _HOST_LAST_HIT.get(host, 0)
        wait = POLITE_DELAY - (time.time() - last)
        _HOST_LAST_HIT[host] = time.time() + max(0, wait)
    if wait > 0:
        time.sleep(wait)


def polite_get(url, stream=False, timeout=REQUEST_TIMEOUT, headers=None):
    """GET a URL with the honest User-Agent and polite spacing.
    Raises ConnectorError with a readable message on failure."""
    _polite_pause(url)
    merged = {"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"}
    if headers:
        merged.update(headers)
    try:
        resp = requests.get(
            url, headers=merged, timeout=timeout, stream=stream
        )
    except requests.RequestException as exc:
        host = urlparse(url).netloc
        raise ConnectorError(
            f"Could not connect to {host} ({exc.__class__.__name__}). "
            "Make sure that the internet connection is up, then try again."
        )
    if resp.status_code >= 400:
        raise ConnectorError(
            f"{urlparse(url).netloc} answered HTTP {resp.status_code} "
            f"for {url.split('?')[0]}"
        )
    return resp


def download_to_temp(url, progress=None, label="file"):
    """Stream a large download to a temporary file.  Returns the path."""
    resp = polite_get(url, stream=True, timeout=DOWNLOAD_TIMEOUT)
    total = to_int(resp.headers.get("Content-Length")) or 0
    handle, tmp_path = tempfile.mkstemp(prefix="lead_engine_")
    done = 0
    try:
        with os.fdopen(handle, "wb") as out:
            for chunk in resp.iter_content(chunk_size=1 << 18):
                out.write(chunk)
                done += len(chunk)
                if progress and total:
                    progress(
                        f"Download of the {label}: "
                        f"{done // 1048576} of {total // 1048576} MB",
                        min(99, int(100 * done / total)),
                    )
    except Exception:
        Path(tmp_path).unlink(missing_ok=True)
        raise
    finally:
        resp.close()
    return Path(tmp_path)


def sniff_delimiter(sample_line):
    """Pick the most frequent delimiter in the first line."""
    best, best_count = ",", 0
    for cand in (",", "\t", "|", ";"):
        count = sample_line.count(cand)
        if count > best_count:
            best, best_count = cand, count
    return best


def iter_tabular_member(zip_path, member_name=None):
    """Open a ZIP and yield rows (lists of strings) from one member.

    The member format is found automatically: .xlsx through openpyxl,
    other files as delimited text with BOM removal and delimiter
    detection.  The first yielded row is the header.
    """
    with zipfile.ZipFile(zip_path) as zf:
        names = [n for n in zf.namelist() if not n.endswith("/")]
        if not names:
            raise ConnectorError("The downloaded ZIP file is empty.")
        if member_name:
            member = None
            for n in names:
                if member_name.lower() in n.lower():
                    member = n
                    break
            if member is None:
                raise ConnectorError(
                    f"The ZIP file has no member with '{member_name}' "
                    "in its name. The source format changed."
                )
        else:
            preferred = [
                n for n in names
                if n.lower().endswith((".xlsx", ".csv", ".tsv", ".txt"))
            ]
            member = preferred[0] if preferred else names[0]

        if member.lower().endswith(".xlsx"):
            import openpyxl

            handle, tmp = tempfile.mkstemp(suffix=".xlsx")
            try:
                with os.fdopen(handle, "wb") as out:
                    out.write(zf.read(member))
                wb = openpyxl.load_workbook(
                    tmp, read_only=True, data_only=True
                )
                try:
                    ws = wb[wb.sheetnames[0]]
                    for row in ws.iter_rows(values_only=True):
                        yield [
                            "" if v is None else str(v).strip() for v in row
                        ]
                finally:
                    wb.close()
            finally:
                Path(tmp).unlink(missing_ok=True)
        else:
            with zf.open(member) as raw:
                text = io.TextIOWrapper(
                    raw, encoding="utf-8-sig", errors="replace"
                )
                first = text.readline()
                if not first:
                    return
                delim = sniff_delimiter(first)
                yield next(csv.reader([first], delimiter=delim))
                for row in csv.reader(text, delimiter=delim):
                    yield row


# ---------------------------------------------------------------------
# Connectors
# ---------------------------------------------------------------------

def _bs(html_text):
    from bs4 import BeautifulSoup

    return BeautifulSoup(html_text, "html.parser")


def _newest_zip_link(page_url, want=None, avoid=None):
    """Scrape a listing page.  Return the newest .zip link whose text or
    href contains all 'want' words and none of the 'avoid' words."""
    resp = polite_get(page_url)
    soup = _bs(resp.text)
    candidates = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href.lower().split("?")[0].endswith(".zip"):
            continue
        # Look only at the file name and the link text.  The folder
        # part of SEC URLs contains the word "exempt" for all files.
        basename = Path(urlparse(href).path).name.lower()
        blob = basename + " " + a.get_text(" ", strip=True).lower()
        if want and not all(w in blob for w in want):
            continue
        if avoid and any(w in blob for w in avoid):
            continue
        full = href if href.startswith("http") else (
            f"https://{urlparse(page_url).netloc}{href}"
            if href.startswith("/")
            else page_url.rstrip("/") + "/" + href
        )
        digits = re.findall(r"(\d{6,8})", Path(urlparse(full).path).name)
        sort_key = digits[-1] if digits else ""
        # File names use MMDDYY.  Change to YYMMDD so that text sort
        # puts the newest date first.
        if len(sort_key) == 6:
            sort_key = sort_key[4:] + sort_key[:4]
        candidates.append((sort_key, full))
    if not candidates:
        return None
    candidates.sort(key=lambda c: c[0], reverse=True)
    return candidates[0][1]


ADV_COLUMNS = {
    "external_id": ["organization crd", "crd number", "crd"],
    "firm_name": ["primary business name", "legal name", "business name"],
    "address": ["main office street address 1", "main office address 1",
                "street address 1"],
    "address2": ["main office street address 2"],
    "city": ["main office city"],
    "state": ["main office state"],
    "zip_code": ["main office postal", "main office zip"],
    "phone": ["main office telephone", "telephone"],
    "website": ["website address", "web site", "website"],
    "aum_dollars": ["total regulatory assets", "5f(2)(c)",
                    "assets under management"],
    "num_employees": ["total number of employees", "5a."],
}


def _run_adv_connector(conn, state, progress, connector_id, want, avoid):
    """Shared logic for the two SEC Form ADV connectors."""
    progress("Search for the newest ZIP on the SEC data page…")
    link = _newest_zip_link(SEC_ADV_PAGE, want=want, avoid=avoid)
    if link is None:
        raise ConnectorError(
            "The SEC Form ADV page shows no matching ZIP link. "
            "The page layout changed — the connector needs an update."
        )
    zip_path = download_to_temp(link, progress, "SEC adviser file")
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    try:
        rows = iter_tabular_member(zip_path)
        header = next(rows, None)
        if header is None:
            raise ConnectorError("The SEC file is empty.")
        colmap = {}
        for field, cands in ADV_COLUMNS.items():
            hit = first_match(header, cands)
            if hit is not None:
                colmap[field] = header.index(hit)
        if "firm_name" not in colmap or "state" not in colmap:
            raise ConnectorError(
                "The SEC file has no recognizable name or state column. "
                "The file format changed."
            )

        def cell(row, field):
            idx = colmap.get(field)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip()

        for row in rows:
            if cell(row, "state").upper() != state:
                continue
            stats["seen"] += 1
            address = cell(row, "address")
            addr2 = cell(row, "address2")
            if addr2:
                address = f"{address}, {addr2}" if address else addr2
            lead = {
                "lead_type": connector_id,
                "external_id": cell(row, "external_id"),
                "firm_name": cell(row, "firm_name"),
                "address": address,
                "city": cell(row, "city"),
                "state": state,
                "zip_code": cell(row, "zip_code"),
                "phone": cell(row, "phone"),
                "website": cell(row, "website"),
                "aum_dollars": to_int(cell(row, "aum_dollars")),
                "num_employees": to_int(cell(row, "num_employees")),
                "source_connector": connector_id,
                "source_detail": f"SEC Form ADV bulk file {Path(urlparse(link).path).name}",
            }
            if not lead["firm_name"]:
                continue
            _, action = upsert_lead(conn, lead)
            stats[action] += 1
            if stats["seen"] % 200 == 0:
                conn.commit()
                progress(f"{stats['seen']} advisers in {state} processed…")
        conn.commit()
    finally:
        zip_path.unlink(missing_ok=True)
    return stats


def run_ria_sec(conn, state, progress):
    return _run_adv_connector(
        conn, state, progress, "ria_sec",
        want=None, avoid=["exempt", "state"],
    )


def run_ria_state(conn, state, progress):
    return _run_adv_connector(
        conn, state, progress, "ria_state",
        want=["state"], avoid=["exempt"],
    )


def run_fdic_banks(conn, state, progress):
    progress("Query of the FDIC BankFind API…")
    params = {
        "filters": f'STALP:"{state}" AND ACTIVE:1',
        "fields": "CERT,NAME,ADDRESS,CITY,STALP,ZIP,WEBADDR",
        "limit": "10000",
        "format": "json",
    }
    url = FDIC_API + "?" + "&".join(
        f"{k}={requests.utils.quote(v)}" for k, v in params.items()
    )
    resp = polite_get(url)
    try:
        payload = resp.json()
    except ValueError:
        raise ConnectorError("The FDIC API did not return JSON.")
    records = payload.get("data", [])
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    for item in records:
        data = item.get("data", item)
        name = str(data.get("NAME", "")).strip()
        if not name:
            continue
        stats["seen"] += 1
        website = str(data.get("WEBADDR", "") or "").strip()
        lead = {
            "lead_type": "bank",
            "external_id": str(data.get("CERT", "")),
            "firm_name": name,
            "address": str(data.get("ADDRESS", "") or ""),
            "city": str(data.get("CITY", "") or ""),
            "state": state,
            "zip_code": str(data.get("ZIP", "") or ""),
            "website": website,
            "source_connector": "fdic_banks",
            "source_detail": f"FDIC BankFind cert #{data.get('CERT', '')}",
        }
        _, action = upsert_lead(conn, lead)
        stats[action] += 1
    conn.commit()
    progress(f"{stats['seen']} FDIC-insured banks in {state}.")
    return stats


NCUA_COLUMNS = {
    "external_id": ["cu_number", "charter"],
    "firm_name": ["cu_name", "name"],
    "address": ["street", "address"],
    "city": ["city"],
    "state": ["state"],
    "zip_code": ["zip"],
    "phone": ["phone"],
    "website": ["web", "url"],
}


def run_ncua_cus(conn, state, progress):
    progress("Search for the newest NCUA quarterly data ZIP…")
    link = _newest_zip_link(NCUA_QUARTERLY_PAGE, want=["call-report"])
    if link is None:
        link = _newest_zip_link(NCUA_QUARTERLY_PAGE)
    if link is None:
        raise ConnectorError(
            "The NCUA quarterly data page shows no ZIP link. "
            "The page layout changed — the connector needs an update."
        )
    zip_path = download_to_temp(link, progress, "NCUA data file")
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    try:
        rows = iter_tabular_member(zip_path, member_name="foicu")
        header = next(rows, None)
        if header is None:
            raise ConnectorError("The NCUA FOICU file is empty.")
        colmap = {}
        for field, cands in NCUA_COLUMNS.items():
            hit = first_match(header, cands)
            if hit is not None:
                colmap[field] = header.index(hit)
        if "firm_name" not in colmap or "state" not in colmap:
            raise ConnectorError(
                "The NCUA FOICU file has no recognizable columns. "
                "The file format changed."
            )

        def cell(row, field):
            idx = colmap.get(field)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip().strip('"')

        for row in rows:
            if cell(row, "state").upper() != state:
                continue
            name = cell(row, "firm_name")
            if not name:
                continue
            stats["seen"] += 1
            if "credit union" not in name.lower():
                name = name.title() + " Credit Union"
            lead = {
                "lead_type": "credit_union",
                "external_id": cell(row, "external_id"),
                "firm_name": name,
                "address": cell(row, "address"),
                "city": cell(row, "city"),
                "state": state,
                "zip_code": cell(row, "zip_code")[:10],
                "phone": cell(row, "phone"),
                "website": cell(row, "website"),
                "source_connector": "ncua_cus",
                "source_detail": (
                    f"NCUA charter #{cell(row, 'external_id')} "
                    f"({Path(urlparse(link).path).name})"
                ),
            }
            _, action = upsert_lead(conn, lead)
            stats[action] += 1
        conn.commit()
    finally:
        zip_path.unlink(missing_ok=True)
    progress(f"{stats['seen']} credit unions in {state}.")
    return stats


PTIN_COLUMNS = {
    "last_name": ["last name", "last_name", "lastname"],
    "first_name": ["first name", "first_name", "firstname"],
    "firm_name": ["business name", "business_name"],
    "address": ["address line 1", "address_line_1", "street", "address1",
                "mailing address"],
    "city": ["city"],
    "state": ["state"],
    "zip_code": ["zip", "postal"],
    "phone": ["phone", "telephone"],
    "email": ["email"],
    "website": ["website", "web site"],
    "external_id": ["ptin"],
}


def run_irs_ptin(conn, state, progress):
    state_name = STATE_NAMES[state].lower()
    progress("Search for the state file link on the IRS PTIN page…")
    resp = polite_get(IRS_PTIN_PAGE)
    soup = _bs(resp.text)
    link = None
    for a in soup.find_all("a", href=True):
        blob = (a["href"] + " " + a.get_text(" ", strip=True)).lower()
        if state_name.replace(" ", "") in blob.replace(" ", "").replace("_", "").replace("-", ""):
            href = a["href"].strip()
            link = href if href.startswith("http") else (
                "https://www.irs.gov" + href
            )
            break
    if link is None:
        raise ConnectorError(
            f"The IRS PTIN page shows no file link for "
            f"{STATE_NAMES[state]}. The page layout changed."
        )
    file_path = download_to_temp(link, progress, "IRS PTIN file")
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    try:
        if zipfile.is_zipfile(file_path):
            rows = iter_tabular_member(file_path)
        else:
            def plain_rows():
                with open(
                    file_path, encoding="utf-8-sig", errors="replace"
                ) as fh:
                    first = fh.readline()
                    if not first:
                        return
                    delim = sniff_delimiter(first)
                    yield next(csv.reader([first], delimiter=delim))
                    for row in csv.reader(fh, delimiter=delim):
                        yield row
            rows = plain_rows()
        header = next(rows, None)
        if header is None:
            raise ConnectorError("The IRS PTIN file is empty.")
        colmap = {}
        for field, cands in PTIN_COLUMNS.items():
            hit = first_match(header, cands)
            if hit is not None:
                colmap[field] = header.index(hit)
        if "last_name" not in colmap and "firm_name" not in colmap:
            raise ConnectorError(
                "The IRS PTIN file has no recognizable name column. "
                "The file format changed."
            )

        def cell(row, field):
            idx = colmap.get(field)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip()

        for row in rows:
            row_state = cell(row, "state").upper()
            if row_state and row_state not in (state, STATE_NAMES[state].upper()):
                continue
            person = " ".join(
                p for p in (cell(row, "first_name"), cell(row, "last_name")) if p
            ).title()
            firm = cell(row, "firm_name") or person
            if not firm:
                continue
            stats["seen"] += 1
            ext = cell(row, "external_id")
            if not ext:
                digest = hashlib.sha1(
                    f"{firm}|{person}|{cell(row, 'zip_code')}".encode()
                ).hexdigest()[:16]
                ext = f"ptin-{digest}"
            lead = {
                "lead_type": "cpa",
                "external_id": ext,
                "firm_name": firm,
                "contact_name": person,
                "address": cell(row, "address"),
                "city": cell(row, "city").title(),
                "state": state,
                "zip_code": cell(row, "zip_code")[:10],
                "phone": cell(row, "phone"),
                "email": cell(row, "email"),
                "website": cell(row, "website"),
                "source_connector": "irs_ptin",
                "source_detail": "IRS PTIN holder file "
                                 f"({STATE_NAMES[state]})",
            }
            _, action = upsert_lead(conn, lead)
            stats[action] += 1
            if stats["seen"] % 500 == 0:
                conn.commit()
                progress(f"{stats['seen']} PTIN holders processed…")
        conn.commit()
    finally:
        file_path.unlink(missing_ok=True)
    progress(f"{stats['seen']} PTIN holders in {state}.")
    return stats


def run_osm_insurance(conn, state, progress):
    progress("Query of the Overpass API for insurance offices…")
    query = f"""
[out:json][timeout:180];
area["ISO3166-2"="US-{state}"][admin_level=4]->.a;
(
  node(area.a)["office"="insurance"];
  way(area.a)["office"="insurance"];
  node(area.a)["shop"="insurance"];
  way(area.a)["shop"="insurance"];
);
out center tags;
"""
    _polite_pause(OVERPASS_API)
    try:
        resp = requests.post(
            OVERPASS_API,
            data={"data": query},
            headers={"User-Agent": USER_AGENT},
            timeout=240,
        )
    except requests.RequestException as exc:
        raise ConnectorError(
            f"Could not connect to the Overpass API "
            f"({exc.__class__.__name__}). Try again later."
        )
    if resp.status_code == 429:
        raise ConnectorError(
            "The Overpass API is busy (HTTP 429). Wait some minutes, "
            "then try again."
        )
    if resp.status_code >= 400:
        raise ConnectorError(
            f"The Overpass API answered HTTP {resp.status_code}."
        )
    try:
        elements = resp.json().get("elements", [])
    except ValueError:
        raise ConnectorError("The Overpass API did not return JSON.")
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    for el in elements:
        tags = el.get("tags", {})
        name = (tags.get("name") or "").strip()
        if not name:
            continue
        stats["seen"] += 1
        house = tags.get("addr:housenumber", "")
        street = tags.get("addr:street", "")
        address = f"{house} {street}".strip()
        lead = {
            "lead_type": "insurance",
            "external_id": f"osm-{el.get('type')}-{el.get('id')}",
            "firm_name": name,
            "address": address,
            "city": tags.get("addr:city", ""),
            "state": state,
            "zip_code": tags.get("addr:postcode", "")[:10],
            "phone": tags.get("phone") or tags.get("contact:phone", ""),
            "email": tags.get("email") or tags.get("contact:email", ""),
            "website": tags.get("website")
                       or tags.get("contact:website", ""),
            "source_connector": "osm_insurance",
            "source_detail": "OpenStreetMap "
                             f"{el.get('type')}/{el.get('id')}",
        }
        _, action = upsert_lead(conn, lead)
        stats[action] += 1
    conn.commit()
    progress(f"{stats['seen']} insurance offices in {state} from OSM.")
    return stats


SBA_COLUMNS = {
    "firm_name": ["borrname", "borrower name", "borr_name"],
    "address": ["borrstreet", "borrower street"],
    "city": ["borrcity", "borrower city"],
    "state": ["borrstate", "borrower state"],
    "zip_code": ["borrzip", "borrower zip"],
    "amount": ["grossapproval", "gross approval"],
    "date": ["approvaldate", "approval date", "approvalfiscalyear"],
    "lender": ["bankname", "thirdpartylender_name", "lender"],
    "lender_city": ["bankcity"],
    "lender_state": ["bankstate"],
    "naics": ["naicscode", "naics"],
}


def run_sba_borrowers(conn, state, progress):
    progress("Ask the SBA open-data catalog for the newest 7(a) file…")
    resp = polite_get(SBA_CKAN_API)
    try:
        payload = resp.json()
    except ValueError:
        raise ConnectorError("The SBA data catalog did not return JSON.")
    resources = payload.get("result", {}).get("resources", [])
    best = None
    for res in resources:
        name = (res.get("name") or "").lower()
        fmt = (res.get("format") or "").lower()
        if "csv" not in fmt:
            continue
        if "7(a)" in name or "7a" in name:
            if "present" in name:
                best = res
                break
            if best is None:
                best = res
    if best is None or not best.get("url"):
        raise ConnectorError(
            "The SBA catalog shows no 7(a) CSV file. "
            "The dataset layout changed."
        )
    url = best["url"]
    progress(f"Stream of {best.get('name', 'the SBA file')}…")
    resp = polite_get(url, stream=True, timeout=DOWNLOAD_TIMEOUT)
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    lender_totals = {}
    try:
        lines = (
            line.decode("utf-8", errors="replace")
            for line in resp.iter_lines()
            if line
        )
        reader = csv.reader(lines)
        header = next(reader, None)
        if header is None:
            raise ConnectorError("The SBA CSV file is empty.")
        colmap = {}
        for field, cands in SBA_COLUMNS.items():
            hit = first_match(header, cands)
            if hit is not None:
                colmap[field] = header.index(hit)
        if "firm_name" not in colmap or "state" not in colmap:
            raise ConnectorError(
                "The SBA CSV has no recognizable borrower columns. "
                "The file format changed."
            )

        def cell(row, field):
            idx = colmap.get(field)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip()

        scanned = 0
        for row in reader:
            scanned += 1
            if scanned % 20000 == 0:
                progress(
                    f"{scanned} loan rows scanned, "
                    f"{stats['seen']} in {state}…"
                )
            if cell(row, "state").upper() != state:
                continue
            name = cell(row, "firm_name")
            if not name:
                continue
            stats["seen"] += 1
            amount = to_int(cell(row, "amount"))
            lender = cell(row, "lender")
            detail_parts = ["SBA 7(a) borrower"]
            if amount:
                detail_parts.append(f"${amount:,}")
            when = cell(row, "date")
            if when:
                detail_parts.append(f"approved {when}")
            if lender:
                detail_parts.append(f"lender: {lender}")
            digest = hashlib.sha1(
                f"{name}|{cell(row, 'zip_code')}|{when}|{amount}".encode()
            ).hexdigest()[:16]
            lead = {
                "lead_type": "borrower",
                "external_id": f"sba-{digest}",
                "firm_name": name.title(),
                "address": cell(row, "address").title(),
                "city": cell(row, "city").title(),
                "state": state,
                "zip_code": cell(row, "zip_code")[:10],
                "source_connector": "sba_borrowers",
                "source_detail": "; ".join(detail_parts),
            }
            _, action = upsert_lead(conn, lead)
            stats[action] += 1
            if lender:
                key = lender.title()
                agg = lender_totals.setdefault(
                    key,
                    {
                        "count": 0,
                        "total": 0,
                        "city": cell(row, "lender_city").title(),
                        "state": cell(row, "lender_state").upper(),
                    },
                )
                agg["count"] += 1
                agg["total"] += amount or 0
            if stats["seen"] % 500 == 0:
                conn.commit()
    finally:
        resp.close()
    stamp = now_iso()
    for name, agg in lender_totals.items():
        conn.execute(
            """INSERT INTO lenders(name, city, state, loan_count,
                total_dollars, last_seen)
               VALUES(?,?,?,?,?,?)
               ON CONFLICT(name, state) DO UPDATE SET
                 loan_count = excluded.loan_count,
                 total_dollars = excluded.total_dollars,
                 city = excluded.city,
                 last_seen = excluded.last_seen""",
            (name, agg["city"], agg["state"], agg["count"], agg["total"],
             stamp),
        )
    conn.commit()
    progress(
        f"{stats['seen']} SBA borrowers in {state}; "
        f"{len(lender_totals)} lenders mapped."
    )
    return stats


CONNECTORS = [
    {
        "id": "ria_sec",
        "label": "SEC-registered advisors",
        "category": "Referral partners",
        "description": "Monthly SEC Form ADV bulk file. Unique key: CRD.",
        "run": run_ria_sec,
        "kind": "auto",
    },
    {
        "id": "ria_state",
        "label": "State-registered advisors",
        "category": "Referral partners",
        "description": "State-registered adviser file from the SEC data page.",
        "run": run_ria_state,
        "kind": "auto",
    },
    {
        "id": "fdic_banks",
        "label": "Community banks (FDIC)",
        "category": "Referral partners",
        "description": "FDIC BankFind Suite API, active banks in the state.",
        "run": run_fdic_banks,
        "kind": "auto",
    },
    {
        "id": "ncua_cus",
        "label": "Credit unions (NCUA)",
        "category": "Referral partners",
        "description": "NCUA quarterly call-report directory (FOICU).",
        "run": run_ncua_cus,
        "kind": "auto",
    },
    {
        "id": "irs_ptin",
        "label": "CPAs / tax preparers (IRS)",
        "category": "Referral partners",
        "description": "IRS PTIN holder FOIA file for the state.",
        "run": run_irs_ptin,
        "kind": "auto",
    },
    {
        "id": "osm_insurance",
        "label": "Insurance agencies (OSM)",
        "category": "Referral partners",
        "description": "OpenStreetMap insurance offices in the state. "
                       "For the paid OCI licensee list, use CSV import.",
        "run": run_osm_insurance,
        "kind": "auto",
    },
    {
        "id": "sba_borrowers",
        "label": "SBA borrowers (FOIA)",
        "category": "Borrower prospects",
        "description": "SBA 7(a) FOIA loans in the state. Also fills the "
                       "lender map.",
        "run": run_sba_borrowers,
        "kind": "auto",
    },
    {
        "id": "import_cre",
        "label": "Commercial RE agents",
        "category": "Referral partners",
        "description": "No good free bulk source. Import a CSV list.",
        "run": None,
        "kind": "import",
        "lead_type": "cre_agent",
    },
    {
        "id": "import_attorney",
        "label": "Business attorneys",
        "category": "Referral partners",
        "description": "No good free bulk source. Import a CSV list.",
        "run": None,
        "kind": "import",
        "lead_type": "attorney",
    },
    {
        "id": "import_oci",
        "label": "Insurance agents (OCI list)",
        "category": "Referral partners",
        "description": "Import a purchased OCI licensee CSV.",
        "run": None,
        "kind": "import",
        "lead_type": "insurance",
    },
    {
        "id": "import_generic",
        "label": "Generic CSV import",
        "category": "Any",
        "description": "Import any CSV with column mapping.",
        "run": None,
        "kind": "import",
        "lead_type": "generic",
    },
]
CONNECTOR_BY_ID = {c["id"]: c for c in CONNECTORS}
AUTO_CONNECTORS = [c for c in CONNECTORS if c["kind"] == "auto"]


# ---------------------------------------------------------------------
# Background jobs
# ---------------------------------------------------------------------

JOBS = {}
JOBS_LOCK = threading.Lock()


def job_snapshot():
    with JOBS_LOCK:
        return {k: dict(v) for k, v in JOBS.items()}


def job_start(key, label):
    """Mark a job as running.  Return False if it already runs."""
    with JOBS_LOCK:
        job = JOBS.get(key)
        if job and job.get("running"):
            return False
        JOBS[key] = {
            "running": True,
            "label": label,
            "message": "Started…",
            "progress": None,
            "error": None,
            "stats": None,
            "finished_at": None,
        }
        return True


def job_update(key, message=None, progress=None):
    with JOBS_LOCK:
        job = JOBS.get(key)
        if not job:
            return
        if message is not None:
            job["message"] = message
        if progress is not None:
            job["progress"] = progress


def job_finish(key, error=None, stats=None, message=None):
    with JOBS_LOCK:
        job = JOBS.get(key)
        if not job:
            return
        job["running"] = False
        job["error"] = error
        job["stats"] = stats
        job["finished_at"] = now_iso()
        if message:
            job["message"] = message
        elif error:
            job["message"] = error


def start_connector_run(connector_id):
    """Run one connector in a background thread.  Returns (ok, msg)."""
    connector = CONNECTOR_BY_ID.get(connector_id)
    if connector is None or connector["kind"] != "auto":
        return False, "Unknown connector."
    key = f"conn:{connector_id}"
    if not job_start(key, connector["label"]):
        return False, "This connector already runs."

    def work():
        conn = get_db()
        started = now_iso()
        run_id = None
        try:
            cur = conn.execute(
                "INSERT INTO runs(connector_id, started_at) VALUES(?,?)",
                (connector_id, started),
            )
            run_id = cur.lastrowid
            conn.commit()
            state = get_target_state(conn)

            def progress(message, pct=None):
                job_update(key, message, pct)

            stats = connector["run"](conn, state, progress)
            progress("Scan for near-duplicate names…")
            run_fuzzy_dedup(conn)
            conn.execute(
                """UPDATE runs SET finished_at = ?, rows_seen = ?,
                    rows_added = ?, rows_updated = ?, rows_deduped = ?
                   WHERE id = ?""",
                (
                    now_iso(), stats["seen"], stats["added"],
                    stats["updated"], stats["deduped"], run_id,
                ),
            )
            conn.commit()
            job_finish(
                key,
                stats=stats,
                message=(
                    f"Done: {stats['seen']} rows, {stats['added']} new, "
                    f"{stats['updated']} updated, "
                    f"{stats['deduped']} matched."
                ),
            )
        except ConnectorError as exc:
            _record_run_error(conn, run_id, str(exc))
            job_finish(key, error=str(exc))
        except Exception as exc:
            msg = (
                f"Unexpected error: {exc.__class__.__name__}: {exc}. "
                "Try again; if it repeats, the source format changed."
            )
            _record_run_error(conn, run_id, msg)
            job_finish(key, error=msg)
        finally:
            conn.close()

    threading.Thread(target=work, daemon=True).start()
    return True, "Started."


def _record_run_error(conn, run_id, message):
    try:
        if run_id is not None:
            conn.execute(
                "UPDATE runs SET finished_at = ?, error = ? WHERE id = ?",
                (now_iso(), message, run_id),
            )
            conn.commit()
    except Exception:
        pass


def start_verify_sweep(lead_ids=None):
    """Verify many leads in a background thread."""
    if not job_start("verify", "Verification sweep"):
        return False, "A verification sweep already runs."

    def work():
        conn = get_db()
        try:
            if lead_ids:
                rows = [
                    conn.execute(
                        "SELECT * FROM leads WHERE id = ?", (i,)
                    ).fetchone()
                    for i in lead_ids
                ]
                rows = [r for r in rows if r is not None]
            else:
                rows = conn.execute("SELECT * FROM leads").fetchall()
            total = len(rows) or 1
            for idx, lead in enumerate(rows, 1):
                job_update(
                    "verify",
                    f"Check {idx} of {total}: {lead['firm_name']}",
                    int(100 * idx / total),
                )
                verify_lead(conn, lead, online=True)
                if idx % 20 == 0:
                    conn.commit()
            conn.commit()
            job_finish(
                "verify", message=f"Done: {total} leads verified."
            )
        except Exception as exc:
            job_finish(
                "verify",
                error=f"The sweep stopped: {exc.__class__.__name__}: {exc}",
            )
        finally:
            conn.close()

    threading.Thread(target=work, daemon=True).start()
    return True, "Started."


def start_contact_scrape(lead_ids, only_unscraped=False):
    """Scrape firm websites for contacts in a background thread."""
    if not job_start("scrape", "Contact scrape"):
        return False, "A contact scrape already runs."

    def work():
        conn = get_db()
        try:
            rows = []
            for lead_id in lead_ids:
                lead = conn.execute(
                    "SELECT * FROM leads WHERE id = ?", (lead_id,)
                ).fetchone()
                if lead is None or not lead["website"]:
                    continue
                if only_unscraped:
                    done = conn.execute(
                        "SELECT 1 FROM scrape_log "
                        "WHERE lead_id = ? AND ok = 1",
                        (lead_id,),
                    ).fetchone()
                    if done:
                        continue
                rows.append(lead)
            total = len(rows) or 1
            found_total = 0
            for idx, lead in enumerate(rows, 1):
                job_update(
                    "scrape",
                    f"Site {idx} of {total}: {lead['firm_name']}",
                    int(100 * idx / total),
                )
                result = scrape_firm(lead["website"])
                stamp = now_iso()
                if result["ok"]:
                    for c in result["contacts"]:
                        dup = conn.execute(
                            "SELECT id FROM contacts WHERE lead_id = ? "
                            "AND lower(name) = lower(?)",
                            (lead["id"], c["name"]),
                        ).fetchone()
                        if dup:
                            continue
                        conn.execute(
                            """INSERT INTO contacts(lead_id, name, title,
                                email, phone, linkedin_url, source_url,
                                confidence, scraped_at)
                               VALUES(?,?,?,?,?,?,?,?,?)""",
                            (
                                lead["id"], c["name"], c["title"],
                                c["email"], c["phone"], c["linkedin_url"],
                                c["source_url"], c["confidence"], stamp,
                            ),
                        )
                    found_total += len(result["contacts"])
                conn.execute(
                    """INSERT INTO scrape_log(lead_id, scraped_at, ok,
                        contacts_found, error) VALUES(?,?,?,?,?)""",
                    (
                        lead["id"], stamp, 1 if result["ok"] else 0,
                        len(result["contacts"]),
                        result["error"] or "",
                    ),
                )
                conn.commit()
            job_finish(
                "scrape",
                message=(
                    f"Done: {total} sites, {found_total} contacts found."
                ),
            )
        except Exception as exc:
            job_finish(
                "scrape",
                error=f"The scrape stopped: {exc.__class__.__name__}: {exc}",
            )
        finally:
            conn.close()

    threading.Thread(target=work, daemon=True).start()
    return True, "Started."


# ---------------------------------------------------------------------
# Google Sheets publish (optional, one-way)
# ---------------------------------------------------------------------

def sheets_available(conn):
    return SERVICE_ACCOUNT_PATH.exists() and bool(
        get_setting(conn, "sheet_url")
    )


def start_publish():
    if not job_start("publish", "Google Sheets publish"):
        return False, "A publish already runs."

    def work():
        conn = get_db()
        try:
            if not SERVICE_ACCOUNT_PATH.exists():
                raise ConnectorError(
                    "service_account.json is not in the app folder."
                )
            sheet_url = get_setting(conn, "sheet_url")
            if not sheet_url:
                raise ConnectorError(
                    "No Sheet URL is set on the Settings page."
                )
            try:
                import gspread
            except ImportError:
                raise ConnectorError(
                    "The gspread package is not installed. "
                    "Run install.bat again."
                )
            job_update("publish", "Connection to Google Sheets…")
            client = gspread.service_account(
                filename=str(SERVICE_ACCOUNT_PATH)
            )
            book = client.open_by_url(sheet_url)
            header = [
                "ID", "Type", "Firm", "Contact", "City", "State", "Phone",
                "Email", "Website", "Score", "Tier", "Status", "Source",
                "Source detail",
            ]
            values = [header]
            for lead in conn.execute(
                "SELECT * FROM all_leads ORDER BY score DESC, firm_name"
            ):
                values.append([
                    lead["id"],
                    LEAD_TYPES.get(lead["lead_type"], lead["lead_type"]),
                    lead["firm_name"], lead["contact_name"],
                    lead["city"], lead["state"], lead["phone"],
                    lead["email"], lead["website"], lead["score"],
                    lead["tier"], lead["status"],
                    lead["source_connector"], lead["source_detail"],
                ])
            job_update(
                "publish", f"Write of {len(values) - 1} rows to the sheet…"
            )
            try:
                sheet = book.worksheet("Leads")
                sheet.clear()
            except Exception:
                sheet = book.add_worksheet(
                    "Leads", rows=len(values) + 10, cols=len(header)
                )
            # One values.update call keeps the app inside the free quota.
            sheet.update(values, "A1")
            set_setting(conn, "last_publish", now_iso())
            conn.commit()
            job_finish(
                "publish",
                message=f"Done: {len(values) - 1} leads published.",
            )
        except ConnectorError as exc:
            job_finish("publish", error=str(exc))
        except Exception as exc:
            job_finish(
                "publish",
                error=f"Publish failed: {exc.__class__.__name__}: {exc}",
            )
        finally:
            conn.close()

    threading.Thread(target=work, daemon=True).start()
    return True, "Started."


# ---------------------------------------------------------------------
# Weekly schedule
# ---------------------------------------------------------------------

SCHEDULER = None


def apply_schedule():
    """Read the schedule settings and set the weekly job."""
    global SCHEDULER
    from apscheduler.schedulers.background import BackgroundScheduler

    conn = get_db()
    try:
        enabled = get_setting(conn, "schedule_enabled", "0") == "1"
        day = get_setting(conn, "schedule_day", "mon")
    finally:
        conn.close()
    if SCHEDULER is None:
        SCHEDULER = BackgroundScheduler(daemon=True)
        SCHEDULER.start()
    for job in SCHEDULER.get_jobs():
        job.remove()
    if enabled:
        SCHEDULER.add_job(
            run_all_connectors_scheduled,
            "cron",
            day_of_week=day,
            hour=6,
            minute=0,
            id="weekly",
        )


def run_all_connectors_scheduled():
    """The weekly job: run each auto connector, one after the other."""
    for connector in AUTO_CONNECTORS:
        ok, _ = start_connector_run(connector["id"])
        if not ok:
            continue
        key = f"conn:{connector['id']}"
        while True:
            time.sleep(5)
            with JOBS_LOCK:
                job = JOBS.get(key)
                if not job or not job.get("running"):
                    break


# ---------------------------------------------------------------------
# Demo data
# ---------------------------------------------------------------------

DEMO_LEADS = [
    ("ria_sec", "100001", "Badger Wealth Management LLC", "Madison",
     "6085550101", "badgerwealth.example.com", "info@badgerwealth.example.com",
     "10 W Mifflin St"),
    ("ria_sec", "100002", "Mendota Capital Advisors", "Madison",
     "6085550102", "mendotacapital.example.com", "", "1 S Pinckney St"),
    ("ria_sec", "100003", "Lakeshore Financial Planning", "Milwaukee",
     "4145550103", "lakeshorefp.example.com", "hello@lakeshorefp.example.com",
     "735 N Water St"),
    ("ria_sec", "100004", "Fox Valley Advisory Group", "Appleton",
     "9205550104", "", "", "100 W College Ave"),
    ("ria_sec", "100005", "Northwoods Asset Management", "Wausau",
     "7155550105", "northwoodsam.example.com", "", ""),
    ("ria_state", "200001", "Kettle Moraine Advisers LLC", "West Bend",
     "2625550106", "", "", "204 S Main St"),
    ("ria_state", "200002", "Driftless Region Planning", "La Crosse",
     "6085550107", "driftlessplanning.example.com", "", "401 Main St"),
    ("bank", "300001", "First Settlers Bank", "Madison", "6085550110",
     "firstsettlers.example.com", "", "1 N Capitol Sq"),
    ("bank", "300002", "Cream City Community Bank", "Milwaukee",
     "4145550111", "creamcitybank.example.com", "", "500 E Wisconsin Ave"),
    ("bank", "300003", "Door County State Bank", "Sturgeon Bay",
     "9205550112", "", "", "20 N 3rd Ave"),
    ("credit_union", "400001", "Blue Harbor Credit Union", "Sheboygan",
     "9205550113", "blueharborcu.example.com", "", "802 N 8th St"),
    ("credit_union", "400002", "Capitol View Credit Union", "Madison",
     "6085550114", "", "", "44 E Main St"),
    ("cpa", "500001", "Hoffman & Reyes CPAs", "Green Bay", "9205550115",
     "hoffmanreyes.example.com", "office@hoffmanreyes.example.com",
     "300 Cherry St"),
    ("cpa", "500002", "Sandra Kowalski Tax Services", "Eau Claire",
     "7155550116", "", "", "212 S Barstow St"),
    ("cpa", "500003", "Riverside Accounting Partners", "Oshkosh",
     "9205550117", "riversideap.example.com", "", "10 Algoma Blvd"),
    ("insurance", "600001", "Harbor Lights Insurance Agency", "Kenosha",
     "2625550118", "harborlightsins.example.com", "", "5710 7th Ave"),
    ("insurance", "600002", "Prairie Shield Insurance", "Sun Prairie",
     "6085550119", "", "", "255 W Main St"),
    ("borrower", "700001", "Third Coast Coffee Roasters", "Milwaukee",
     "4145550120", "thirdcoastcoffee.example.com", "", "168 N Broadway"),
    ("borrower", "700002", "Granite Peak Fabrication LLC", "Wausau",
     "7155550121", "", "", "1200 Cleveland Ave"),
    ("borrower", "700003", "Lakeside Orthodontics", "Madison",
     "6085550122", "", "", "702 N Midvale Blvd"),
    ("borrower", "700004", "Old World Cheese Haus", "New Glarus",
     "6085550123", "oldworldcheese.example.com", "", "534 1st St"),
    ("attorney", "800001", "Vandenberg Business Law", "Green Bay",
     "9205550124", "vandenberglaw.example.com", "", "125 S Jefferson St"),
    ("cre_agent", "900001", "Meridian Commercial Realty", "Brookfield",
     "2625550125", "meridiancre.example.com", "", "200 S Executive Dr"),
    ("generic", "", "Great Lakes Equipment Finance", "Racine",
     "2625550126", "", "", ""),
]

DEMO_CONNECTOR = {
    "ria_sec": "ria_sec", "ria_state": "ria_state", "bank": "fdic_banks",
    "credit_union": "ncua_cus", "cpa": "irs_ptin",
    "insurance": "osm_insurance", "borrower": "sba_borrowers",
    "attorney": "csv_import", "cre_agent": "csv_import",
    "generic": "csv_import",
}


def seed_demo():
    """Put demo data in the database for a safe offline test."""
    conn = get_db()
    try:
        count = conn.execute(
            "SELECT COUNT(*) AS n FROM leads"
        ).fetchone()["n"]
        if count > 0:
            return
        for (ltype, ext, name, city, phone, site, email, addr) in DEMO_LEADS:
            upsert_lead(conn, {
                "lead_type": ltype,
                "external_id": ext,
                "firm_name": name,
                "address": addr,
                "city": city,
                "state": "WI",
                "zip_code": "53703" if city == "Madison" else "53202",
                "phone": phone,
                "email": email,
                "website": site,
                "source_connector": DEMO_CONNECTOR[ltype],
                "source_detail": "Demo seed row",
            })
        # Two near-duplicate rows for the dedup review queue.  A direct
        # INSERT bypasses the exact-match cascade on purpose.
        stamp = now_iso()
        for name, ext in (
            ("Acme Financial LLC", "100099"),
            ("Acme Financial", ""),
        ):
            conn.execute(
                """INSERT INTO leads(lead_type, external_id, firm_name,
                    city, state, zip_code, source_connector,
                    source_detail, first_seen, last_seen)
                   VALUES('ria_sec', ?, ?, 'Madison', 'WI', '53703',
                    'ria_sec', 'Demo duplicate pair', ?, ?)""",
                (ext, name, stamp, stamp),
            )
        demo_contacts = [
            ("Badger Wealth Management LLC", "Maria Lindqvist",
             "Managing Partner", "maria@badgerwealth.example.com"),
            ("Badger Wealth Management LLC", "Tom Okafor",
             "Senior Advisor", "tom@badgerwealth.example.com"),
            ("Hoffman & Reyes CPAs", "Elena Reyes", "Partner, CPA",
             "elena@hoffmanreyes.example.com"),
            ("First Settlers Bank", "Dale Vang",
             "VP Commercial Lending", ""),
        ]
        for firm, cname, title, cemail in demo_contacts:
            lead = conn.execute(
                "SELECT id FROM leads WHERE firm_name = ?", (firm,)
            ).fetchone()
            if lead:
                conn.execute(
                    """INSERT INTO contacts(lead_id, name, title, email,
                        confidence, scraped_at, source_url)
                       VALUES(?,?,?,?,0.85,?, 'demo')""",
                    (lead["id"], cname, title, cemail, stamp),
                )
        conn.commit()
        # Offline checks only: no network traffic in demo mode.
        for lead in conn.execute("SELECT * FROM leads").fetchall():
            verify_lead(conn, lead, online=False)
        conn.commit()
        run_fuzzy_dedup(conn)
    finally:
        conn.close()


# ---------------------------------------------------------------------
# Flask app and templates
# ---------------------------------------------------------------------

app = Flask(__name__)

BASE_HTML = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Lead Engine v2</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body { font-family: ui-sans-serif, system-ui, sans-serif; }
  .tier-A { background:#dcfce7; color:#166534; }
  .tier-B { background:#fef9c3; color:#854d0e; }
  .tier-C { background:#fee2e2; color:#991b1b; }
  .badge { display:inline-block; padding:0 .5rem; border-radius:9999px;
           font-size:.75rem; font-weight:600; line-height:1.25rem; }
  table.data td, table.data th { padding:.4rem .6rem; }
</style>
</head>
<body class="bg-slate-100 text-slate-800">
<header class="bg-slate-900 text-white">
  <div class="max-w-7xl mx-auto px-4 py-3 flex items-center gap-6">
    <span class="text-lg font-bold tracking-tight">Lead Engine
      <span class="text-sky-400">v2</span></span>
    <nav class="flex gap-4 text-sm">
      <a href="{{ url_for('page_dashboard') }}"
         class="hover:text-sky-300">Dashboard</a>
      <a href="{{ url_for('page_leads') }}"
         class="hover:text-sky-300">Leads</a>
      <a href="{{ url_for('page_dedup') }}" class="hover:text-sky-300">
        Dedup Review
        {% if dedup_count %}<span class="badge bg-amber-500 text-white">
          {{ dedup_count }}</span>{% endif %}</a>
      <a href="{{ url_for('page_import') }}"
         class="hover:text-sky-300">Import</a>
      <a href="{{ url_for('page_lenders') }}"
         class="hover:text-sky-300">Lenders</a>
      <a href="{{ url_for('page_settings') }}"
         class="hover:text-sky-300">Settings</a>
    </nav>
    <span class="flex-1"></span>
    <span id="globalStatus" class="text-xs text-slate-300"></span>
    <button onclick="quitApp()"
      class="text-xs bg-rose-600 hover:bg-rose-500 px-3 py-1 rounded">
      Quit App</button>
  </div>
</header>
{% if request.args.get('msg') %}
<div class="max-w-7xl mx-auto px-4 pt-3">
  <div class="bg-sky-50 border border-sky-300 text-sky-900 text-sm
              rounded px-3 py-2">{{ request.args.get('msg') }}</div>
</div>
{% endif %}
<main class="max-w-7xl mx-auto px-4 py-6">
{% block body %}{% endblock %}
</main>
<script>
function quitApp() {
  if (!confirm('Stop the Lead Engine server?')) return;
  fetch('/quit', {method: 'POST'}).finally(function () {
    document.body.innerHTML =
      '<div style="padding:4rem;text-align:center;font-family:sans-serif">' +
      '<h1>Lead Engine stopped.</h1><p>You can close this tab.</p></div>';
  });
}
function postAction(url, body) {
  return fetch(url, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(body || {})
  }).then(function (r) { return r.json(); });
}
</script>
{% block scripts %}{% endblock %}
</body>
</html>"""

DASHBOARD_HTML = """{% extends "base.html" %}
{% block body %}
<div class="grid grid-cols-2 md:grid-cols-5 gap-4 mb-6">
  <div class="bg-white rounded-lg shadow p-4">
    <div class="text-3xl font-bold">{{ total }}</div>
    <div class="text-sm text-slate-500">Total leads</div></div>
  <div class="bg-white rounded-lg shadow p-4">
    <div class="text-3xl font-bold text-green-700">{{ tiers.get('A', 0) }}</div>
    <div class="text-sm text-slate-500">Tier A (75+)</div></div>
  <div class="bg-white rounded-lg shadow p-4">
    <div class="text-3xl font-bold text-yellow-700">{{ tiers.get('B', 0) }}</div>
    <div class="text-sm text-slate-500">Tier B (50–74)</div></div>
  <div class="bg-white rounded-lg shadow p-4">
    <div class="text-3xl font-bold text-red-700">{{ tiers.get('C', 0) }}</div>
    <div class="text-sm text-slate-500">Tier C (&lt;50)</div></div>
  <div class="bg-white rounded-lg shadow p-4">
    <div class="text-3xl font-bold">{{ contact_total }}</div>
    <div class="text-sm text-slate-500">Named contacts</div></div>
</div>

<div class="grid md:grid-cols-3 gap-6">
<div class="md:col-span-1">
  <h2 class="font-semibold mb-2">Leads by type</h2>
  <div class="bg-white rounded-lg shadow divide-y divide-slate-100">
  {% for row in by_type %}
    <div class="flex justify-between px-4 py-2 text-sm">
      <a class="text-sky-700 hover:underline"
         href="{{ url_for('page_leads', lead_type=row['lead_type']) }}">
         {{ type_labels.get(row['lead_type'], row['lead_type']) }}</a>
      <span class="font-semibold">{{ row['n'] }}</span>
    </div>
  {% else %}
    <div class="px-4 py-3 text-sm text-slate-500">
      No leads yet. Run a connector to start.</div>
  {% endfor %}
  </div>
</div>

<div class="md:col-span-2">
  <h2 class="font-semibold mb-2">Source connectors
    <span class="text-xs font-normal text-slate-500">
      (target state: {{ target_state }})</span></h2>
  <div class="grid sm:grid-cols-2 gap-4">
  {% for c in connectors %}
    <div class="bg-white rounded-lg shadow p-4 flex flex-col gap-2">
      <div class="flex items-start justify-between gap-2">
        <div>
          <div class="font-semibold text-sm">{{ c.label }}</div>
          <div class="text-xs text-slate-500">{{ c.description }}</div>
        </div>
        {% if c.kind == 'auto' %}
        <button data-conn="{{ c.id }}" onclick="runConnector('{{ c.id }}')"
          class="run-btn shrink-0 text-xs bg-sky-600 hover:bg-sky-500
                 text-white px-3 py-1.5 rounded">Run Now</button>
        {% else %}
        <a href="{{ url_for('page_import', type=c.lead_type) }}"
          class="shrink-0 text-xs bg-slate-600 hover:bg-slate-500
                 text-white px-3 py-1.5 rounded">CSV Import</a>
        {% endif %}
      </div>
      <div class="text-xs text-slate-600 min-h-4"
           id="status-{{ c.id }}">
        {% set run = last_runs.get(c.id) %}
        {% if run %}
          {% if run['error'] %}
            <span class="text-rose-700">Last run failed:
              {{ run['error'] }}</span>
          {% else %}
            Last run {{ run['finished_at'] }}:
            {{ run['rows_seen'] }} rows, {{ run['rows_added'] }} new,
            {{ run['rows_updated'] }} updated,
            {{ run['rows_deduped'] }} matched.
          {% endif %}
        {% endif %}
      </div>
    </div>
  {% endfor %}
  </div>
</div>
</div>
{% endblock %}
{% block scripts %}
<script>
function runConnector(id) {
  var btn = document.querySelector('[data-conn="' + id + '"]');
  if (btn) { btn.disabled = true; btn.textContent = 'Running…'; }
  postAction('/run/' + id).then(function (r) {
    if (!r.ok) {
      alert(r.message);
      if (btn) { btn.disabled = false; btn.textContent = 'Run Now'; }
    }
  });
}
function refresh() {
  fetch('/status').then(function (r) { return r.json(); })
  .then(function (s) {
    var running = [];
    Object.keys(s.jobs).forEach(function (key) {
      var job = s.jobs[key];
      if (key.indexOf('conn:') === 0) {
        var id = key.slice(5);
        var el = document.getElementById('status-' + id);
        var btn = document.querySelector('[data-conn="' + id + '"]');
        if (job.running) {
          if (el) el.innerHTML = '<span class="text-sky-700">' +
            job.message + (job.progress !== null
              ? ' (' + job.progress + '%)' : '') + '</span>';
          if (btn) { btn.disabled = true; btn.textContent = 'Running…'; }
        } else {
          if (btn) { btn.disabled = false; btn.textContent = 'Run Now'; }
          if (el && job.finished_at) {
            el.innerHTML = job.error
              ? '<span class="text-rose-700">' + job.error + '</span>'
              : '<span class="text-green-700">' + job.message + '</span>';
          }
        }
      }
      if (job.running) running.push(job.label + ': ' + job.message);
    });
    document.getElementById('globalStatus').textContent =
      running.join('  |  ');
  }).catch(function () {});
}
setInterval(refresh, 1500);
refresh();
</script>
{% endblock %}"""

LEADS_HTML = """{% extends "base.html" %}
{% block body %}
<div class="flex items-center justify-between mb-4">
  <h1 class="text-xl font-bold">Leads
    <span class="text-sm font-normal text-slate-500">
      ({{ result_count }} match the filters)</span></h1>
</div>
<form method="get" class="bg-white rounded-lg shadow p-3 mb-4 flex
      flex-wrap gap-3 items-end text-sm">
  <label class="flex flex-col gap-1">Type
    <select name="lead_type" class="border rounded px-2 py-1">
      <option value="">All</option>
      {% for key, label in type_labels.items() %}
      <option value="{{ key }}"
        {% if request.args.get('lead_type') == key %}selected{% endif %}>
        {{ label }}</option>
      {% endfor %}
    </select></label>
  <label class="flex flex-col gap-1">Tier
    <select name="tier" class="border rounded px-2 py-1">
      <option value="">All</option>
      {% for t in ['A', 'B', 'C'] %}
      <option value="{{ t }}"
        {% if request.args.get('tier') == t %}selected{% endif %}>
        {{ t }}</option>
      {% endfor %}
    </select></label>
  <label class="flex flex-col gap-1">State
    <input name="state" value="{{ request.args.get('state', '') }}"
      size="4" class="border rounded px-2 py-1"></label>
  <label class="flex flex-col gap-1">Search
    <input name="q" value="{{ request.args.get('q', '') }}"
      placeholder="name, city, phone, email…"
      class="border rounded px-2 py-1 w-56"></label>
  <button class="bg-sky-600 hover:bg-sky-500 text-white px-4 py-1.5
    rounded">Filter</button>
  <a href="{{ url_for('page_leads') }}"
     class="text-slate-500 hover:underline">Reset</a>
  <span class="flex-1"></span>
  <a href="{{ url_for('export_leads') }}?{{ request.query_string.decode() }}"
     class="bg-emerald-600 hover:bg-emerald-500 text-white px-4 py-1.5
     rounded">Export CSV</a>
  <a href="{{ url_for('export_contacts') }}?{{ request.query_string.decode() }}"
     class="bg-emerald-700 hover:bg-emerald-600 text-white px-4 py-1.5
     rounded">Export contacts</a>
</form>

<div class="bg-white rounded-lg shadow p-3 mb-4 flex flex-wrap gap-3
     items-center text-sm">
  <span class="font-semibold">Bulk actions on the filtered set:</span>
  <button onclick="bulkVerify()" id="btnVerify"
    class="bg-indigo-600 hover:bg-indigo-500 text-white px-3 py-1.5
    rounded">Verify all</button>
  <button onclick="bulkScrape()" id="btnScrape"
    class="bg-violet-600 hover:bg-violet-500 text-white px-3 py-1.5
    rounded">Find contacts</button>
  <label class="flex items-center gap-1">
    <input type="checkbox" id="onlyUnscraped" checked>
    only sites without a completed scrape</label>
  <span id="bulkStatus" class="text-xs text-slate-600"></span>
</div>

<div class="bg-white rounded-lg shadow overflow-x-auto">
<table class="data w-full text-sm">
<thead class="bg-slate-50 text-left text-xs uppercase text-slate-500">
<tr>
  {% for col, label in [('firm_name', 'Firm'), ('lead_type', 'Type'),
       ('city', 'City'), ('state', 'St'), ('phone', 'Phone'),
       ('score', 'Score'), ('tier', 'Tier'),
       ('contact_count', 'Contacts'), ('last_seen', 'Last seen')] %}
  <th><a class="hover:underline"
    href="{{ sort_url(col) }}">{{ label }}
    {%- if request.args.get('sort', 'score') == col %}
      {{ '▴' if request.args.get('dir') == 'asc' else '▾' }}
    {%- endif %}</a></th>
  {% endfor %}
</tr>
</thead>
<tbody class="divide-y divide-slate-100">
{% for lead in leads %}
<tr class="hover:bg-sky-50 cursor-pointer"
    onclick="openDetail({{ lead['id'] }})">
  <td class="font-medium">{{ lead['firm_name'] }}
    {% if lead['contact_name'] %}<span class="text-xs text-slate-500">
      · {{ lead['contact_name'] }}</span>{% endif %}</td>
  <td class="text-xs">{{ type_labels.get(lead['lead_type'],
    lead['lead_type']) }}</td>
  <td>{{ lead['city'] }}</td>
  <td>{{ lead['state'] }}</td>
  <td class="whitespace-nowrap">{{ lead['phone'] }}</td>
  <td class="font-semibold">{{ lead['score'] }}</td>
  <td><span class="badge tier-{{ lead['tier'] }}">{{ lead['tier'] }}</span></td>
  <td>{{ lead['contact_count'] }}</td>
  <td class="text-xs text-slate-500 whitespace-nowrap">
    {{ lead['last_seen'][:10] }}</td>
</tr>
{% else %}
<tr><td colspan="9" class="text-center text-slate-500 py-8">
  No leads match the filters.</td></tr>
{% endfor %}
</tbody>
</table>
</div>
{% if result_count > leads|length %}
<p class="text-xs text-slate-500 mt-2">The table shows the first
{{ leads|length }} rows. Make the filters narrower, or use Export CSV
to get all {{ result_count }} rows.</p>
{% endif %}
<p class="text-xs text-slate-500 mt-4">{{ compliance_note }}</p>

<div id="detailPanel" class="fixed top-0 right-0 h-full w-full sm:w-[480px]
     bg-white shadow-2xl border-l border-slate-200 overflow-y-auto hidden
     z-50"></div>
{% endblock %}
{% block scripts %}
<script>
var currentFilters = window.location.search.slice(1);
function bulkVerify() {
  document.getElementById('btnVerify').disabled = true;
  postAction('/verify/start?' + currentFilters).then(function (r) {
    if (!r.ok) alert(r.message);
  });
}
function bulkScrape() {
  document.getElementById('btnScrape').disabled = true;
  var only = document.getElementById('onlyUnscraped').checked;
  postAction('/scrape/bulk?' + currentFilters,
             {only_unscraped: only}).then(function (r) {
    if (!r.ok) alert(r.message);
  });
}
function pollJobs() {
  fetch('/status').then(function (r) { return r.json(); })
  .then(function (s) {
    var parts = [];
    ['verify', 'scrape'].forEach(function (k) {
      var job = s.jobs[k];
      if (!job) return;
      var btn = document.getElementById(
        k === 'verify' ? 'btnVerify' : 'btnScrape');
      if (job.running) {
        parts.push(job.label + ': ' + job.message +
          (job.progress !== null ? ' (' + job.progress + '%)' : ''));
        if (btn) btn.disabled = true;
      } else {
        if (btn) btn.disabled = false;
        if (job.finished_at) parts.push(job.label + ': ' +
          (job.error || job.message));
      }
    });
    document.getElementById('bulkStatus').textContent = parts.join(' | ');
  }).catch(function () {});
}
setInterval(pollJobs, 2000);
pollJobs();

function openDetail(id) {
  var panel = document.getElementById('detailPanel');
  panel.classList.remove('hidden');
  panel.innerHTML = '<div class="p-6 text-sm text-slate-500">Loading…</div>';
  fetch('/api/lead/' + id).then(function (r) { return r.json(); })
  .then(function (d) { renderDetail(d); });
}
function closeDetail() {
  document.getElementById('detailPanel').classList.add('hidden');
}
function esc(s) {
  return String(s == null ? '' : s).replace(/[&<>"]/g, function (c) {
    return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];
  });
}
function renderDetail(d) {
  var l = d.lead;
  var html = '<div class="p-5 text-sm">' +
    '<div class="flex justify-between items-start mb-3">' +
    '<div><h2 class="text-lg font-bold">' + esc(l.firm_name) + '</h2>' +
    '<div class="text-slate-500">' + esc(d.type_label) +
    (l.external_id ? ' · ID ' + esc(l.external_id) : '') + '</div></div>' +
    '<button onclick="closeDetail()" class="text-slate-400 ' +
    'hover:text-slate-700 text-xl leading-none">&times;</button></div>' +
    '<div class="mb-3"><span class="badge tier-' + esc(l.tier) + '">Tier ' +
    esc(l.tier) + '</span> <span class="font-bold">' + l.score +
    '/100</span></div>';

  html += '<dl class="grid grid-cols-3 gap-y-1 mb-4">';
  [['Contact', l.contact_name], ['Address', l.address],
   ['City', l.city + (l.state ? ', ' + l.state : '') + ' ' + l.zip_code],
   ['Phone', l.phone], ['Email', l.email], ['Website', l.website],
   ['AUM', l.aum_dollars ? '$' + Number(l.aum_dollars).toLocaleString() : ''],
   ['Employees', l.num_employees], ['Source', l.source_connector],
   ['Detail', l.source_detail],
   ['Also seen in', l.corroborators], ['Status', l.status],
   ['First seen', l.first_seen], ['Notes', l.notes]]
  .forEach(function (pair) {
    if (!pair[1]) return;
    html += '<dt class="text-slate-500">' + esc(pair[0]) +
      '</dt><dd class="col-span-2">' + esc(pair[1]) + '</dd>';
  });
  html += '</dl>';

  html += '<div class="flex gap-2 mb-4">' +
    '<button onclick="verifyLead(' + l.id + ')" class="bg-indigo-600 ' +
    'text-white text-xs px-3 py-1.5 rounded hover:bg-indigo-500">' +
    'Verify now</button>' +
    '<button onclick="scrapeLead(' + l.id + ')" class="bg-violet-600 ' +
    'text-white text-xs px-3 py-1.5 rounded hover:bg-violet-500">' +
    'Find Contacts</button>' +
    '<span id="detailMsg" class="text-xs text-slate-500 self-center"></span>' +
    '</div>';

  html += '<h3 class="font-semibold mb-1">Why this score</h3>' +
    '<table class="data w-full text-xs mb-4 border border-slate-100">' +
    '<tr class="bg-slate-50"><th class="text-left">Signal</th>' +
    '<th>Weight</th><th>Result</th><th class="text-left">Detail</th></tr>';
  d.breakdown.forEach(function (b) {
    html += '<tr class="border-t border-slate-100"><td>' + esc(b.label) +
      '</td><td class="text-center">' + b.weight +
      '</td><td class="text-center">' +
      (b.passed ? '<span class="text-green-700">✔</span>'
                : '<span class="text-rose-700">✘</span>') +
      '</td><td>' + esc(b.detail) +
      (b.checked_at ? '<div class="text-slate-400">' + esc(b.checked_at) +
       '</div>' : '') + '</td></tr>';
  });
  html += '</table>';

  html += '<h3 class="font-semibold mb-1">Contacts (' +
    d.contacts.length + ')</h3>';
  if (d.contacts.length === 0) {
    html += '<p class="text-xs text-slate-500 mb-3">No contacts yet. ' +
      'Use Find Contacts.</p>';
  } else {
    html += '<div class="divide-y divide-slate-100 mb-3">';
    d.contacts.forEach(function (c) {
      html += '<div class="py-2"><div class="font-medium">' + esc(c.name) +
        (c.title ? ' <span class="text-slate-500 font-normal">— ' +
         esc(c.title) + '</span>' : '') + '</div>' +
        '<div class="text-xs text-slate-600">' +
        [c.email, c.phone].filter(Boolean).map(esc).join(' · ') +
        (c.linkedin_url ? ' · <a class="text-sky-700 underline" href="' +
         esc(c.linkedin_url) + '" target="_blank">LinkedIn</a>' : '') +
        ' <span class="text-slate-400">(confidence ' + c.confidence +
        ')</span></div></div>';
    });
    html += '</div>';
  }
  html += '</div>';
  document.getElementById('detailPanel').innerHTML = html;
}
function verifyLead(id) {
  document.getElementById('detailMsg').textContent = 'Checks run…';
  postAction('/api/lead/' + id + '/verify').then(function () {
    openDetail(id);
  });
}
function scrapeLead(id) {
  document.getElementById('detailMsg').textContent =
    'Scrape started. Watch the status above.';
  postAction('/api/lead/' + id + '/scrape').then(function (r) {
    if (!r.ok) alert(r.message);
  });
}
</script>
{% endblock %}"""

DEDUP_HTML = """{% extends "base.html" %}
{% block body %}
<div class="flex items-center justify-between mb-4">
  <h1 class="text-xl font-bold">Dedup Review
    <span class="text-sm font-normal text-slate-500">
      ({{ pairs|length }} pairs wait for review)</span></h1>
  <button onclick="postAction('/dedup/scan').then(function(){location.reload()})"
    class="text-xs bg-sky-600 hover:bg-sky-500 text-white px-3 py-1.5
    rounded">Scan for duplicates now</button>
</div>
{% for pair in pairs %}
<div class="bg-white rounded-lg shadow mb-4 p-4">
  <div class="text-xs text-slate-500 mb-2">{{ pair.reason }}
    (similarity {{ pair.similarity }}%)</div>
  <div class="grid md:grid-cols-2 gap-4">
    {% for lead in [pair.a, pair.b] %}
    <div class="border rounded p-3 text-sm
         {{ 'border-emerald-400' if lead['id'] == pair.keep_id
            else 'border-slate-200' }}">
      <div class="font-semibold">{{ lead['firm_name'] }}
        {% if lead['id'] == pair.keep_id %}
        <span class="text-xs text-emerald-700">(richer record — kept on
          merge)</span>{% endif %}</div>
      <dl class="grid grid-cols-3 gap-y-0.5 mt-1 text-xs">
        {% for label, value in [
            ('Type', type_labels.get(lead['lead_type'], lead['lead_type'])),
            ('External ID', lead['external_id']),
            ('Address', lead['address']), ('City', lead['city']),
            ('Phone', lead['phone']), ('Email', lead['email']),
            ('Website', lead['website']), ('Source', lead['source_connector']),
            ('Score', lead['score'])] %}
        <dt class="text-slate-500">{{ label }}</dt>
        <dd class="col-span-2">{{ value if value not in (None, '') else '—' }}</dd>
        {% endfor %}
      </dl>
    </div>
    {% endfor %}
  </div>
  <div class="flex gap-2 mt-3">
    <button onclick="resolvePair({{ pair.id }}, 'merge')"
      class="bg-emerald-600 hover:bg-emerald-500 text-white text-xs px-4
      py-1.5 rounded">Merge</button>
    <button onclick="resolvePair({{ pair.id }}, 'distinct')"
      class="bg-slate-500 hover:bg-slate-400 text-white text-xs px-4
      py-1.5 rounded">Not a duplicate</button>
  </div>
</div>
{% else %}
<div class="bg-white rounded-lg shadow p-8 text-center text-slate-500">
  The review queue is empty. Good.</div>
{% endfor %}
{% endblock %}
{% block scripts %}
<script>
function resolvePair(id, action) {
  postAction('/dedup/' + id + '/' + action).then(function (r) {
    if (r.ok) location.reload(); else alert(r.message);
  });
}
</script>
{% endblock %}"""

IMPORT_HTML = """{% extends "base.html" %}
{% block body %}
<h1 class="text-xl font-bold mb-4">CSV Import</h1>
<div class="bg-white rounded-lg shadow p-5 max-w-2xl">
<form method="post" action="{{ url_for('import_upload') }}"
      enctype="multipart/form-data" class="flex flex-col gap-4 text-sm">
  <label class="flex flex-col gap-1 font-medium">Lead type of the rows
    <select name="lead_type" class="border rounded px-2 py-1.5">
      {% for key, label in type_labels.items() %}
      <option value="{{ key }}"
        {% if preselect == key %}selected{% endif %}>{{ label }}</option>
      {% endfor %}
    </select></label>
  <label class="flex flex-col gap-1 font-medium">CSV file
    <input type="file" name="file" accept=".csv,.txt,.tsv" required
      class="border rounded px-2 py-1.5"></label>
  <p class="text-xs text-slate-500">The first row must hold the column
    names. On the next page, you map your columns to lead fields and see
    a preview of the first 10 mapped rows before the commit.</p>
  <button class="bg-sky-600 hover:bg-sky-500 text-white px-4 py-2
    rounded self-start">Upload and map columns</button>
</form>
</div>
{% endblock %}"""

MAPPING_HTML = """{% extends "base.html" %}
{% block body %}
<h1 class="text-xl font-bold mb-1">Map the columns</h1>
<p class="text-sm text-slate-500 mb-4">{{ filename }} —
  {{ row_count }} data rows found. Set a source column for each lead
  field that you want to fill.</p>
<form method="post" action="{{ url_for('import_commit') }}">
<input type="hidden" name="token" value="{{ token }}">
<input type="hidden" name="lead_type" value="{{ lead_type }}">
<div class="grid md:grid-cols-3 gap-x-6 gap-y-2 bg-white rounded-lg
     shadow p-4 mb-4 text-sm">
  {% for field, label in import_fields %}
  <label class="flex items-center justify-between gap-2">{{ label }}
    <select name="map_{{ field }}" onchange="preview()"
      class="border rounded px-1 py-0.5 text-xs w-44 map-select"
      data-field="{{ field }}">
      <option value="">— skip —</option>
      {% for col in columns %}
      <option value="{{ loop.index0 }}"
        {% if guesses.get(field) == loop.index0 %}selected{% endif %}>
        {{ col }}</option>
      {% endfor %}
    </select></label>
  {% endfor %}
</div>
<h2 class="font-semibold text-sm mb-1">Preview of the first 10 mapped
  rows</h2>
<div class="bg-white rounded-lg shadow overflow-x-auto mb-4">
  <table class="data w-full text-xs" id="previewTable"></table>
</div>
<button class="bg-emerald-600 hover:bg-emerald-500 text-white px-5 py-2
  rounded">Commit import</button>
<a href="{{ url_for('page_import') }}"
   class="ml-3 text-slate-500 hover:underline text-sm">Cancel</a>
</form>
{% endblock %}
{% block scripts %}
<script>
var SAMPLE = {{ sample|tojson }};
var FIELDS = {{ field_list|tojson }};
function esc(s) {
  return String(s == null ? '' : s).replace(/[&<>"]/g, function (c) {
    return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c];
  });
}
function preview() {
  var mapping = {};
  document.querySelectorAll('.map-select').forEach(function (sel) {
    if (sel.value !== '') mapping[sel.dataset.field] = parseInt(sel.value);
  });
  var head = '<tr class="bg-slate-50">';
  FIELDS.forEach(function (f) {
    if (f[0] in mapping) head += '<th class="text-left">' + esc(f[1]) +
      '</th>';
  });
  head += '</tr>';
  var body = '';
  SAMPLE.forEach(function (row) {
    body += '<tr class="border-t border-slate-100">';
    FIELDS.forEach(function (f) {
      if (f[0] in mapping) {
        body += '<td>' + esc(row[mapping[f[0]]] || '') + '</td>';
      }
    });
    body += '</tr>';
  });
  document.getElementById('previewTable').innerHTML = head + body;
}
preview();
</script>
{% endblock %}"""

SETTINGS_HTML = """{% extends "base.html" %}
{% block body %}
<h1 class="text-xl font-bold mb-4">Settings</h1>
<div class="grid md:grid-cols-2 gap-6 items-start">
<form method="post" class="bg-white rounded-lg shadow p-5 text-sm
      flex flex-col gap-4">
  <h2 class="font-semibold">Score weights</h2>
  <p class="text-xs text-slate-500 -mt-3">The score of a lead is the sum
    of the weights of the passed checks, scaled to 100. After a change,
    the app computes all scores again.</p>
  <div class="grid grid-cols-2 gap-3">
  {% for key, label in weight_labels.items() %}
    <label class="flex items-center justify-between gap-2">{{ label }}
      <input type="number" min="0" max="100" name="{{ key }}"
        value="{{ weights[key] }}"
        class="border rounded px-2 py-1 w-20"></label>
  {% endfor %}
  </div>
  <h2 class="font-semibold">Target state</h2>
  <select name="target_state" class="border rounded px-2 py-1.5 w-40">
    {% for s in supported_states %}
    <option value="{{ s }}" {% if s == target_state %}selected{% endif %}>
      {{ s }} — {{ state_names[s] }}</option>
    {% endfor %}
  </select>
  <h2 class="font-semibold">Weekly schedule</h2>
  <label class="flex items-center gap-2">
    <input type="checkbox" name="schedule_enabled" value="1"
      {% if schedule_enabled %}checked{% endif %}>
    Run all connectors automatically each week</label>
  <label class="flex items-center gap-2">Day
    <select name="schedule_day" class="border rounded px-2 py-1">
      {% for day, label in [('mon', 'Monday'), ('tue', 'Tuesday'),
          ('wed', 'Wednesday'), ('thu', 'Thursday'), ('fri', 'Friday'),
          ('sat', 'Saturday'), ('sun', 'Sunday')] %}
      <option value="{{ day }}"
        {% if day == schedule_day %}selected{% endif %}>{{ label }}</option>
      {% endfor %}
    </select> at 06:00</label>
  <p class="text-xs text-slate-500">NOTE: The SEC updates the Form ADV
    data monthly. A weekly run is the practical maximum. More frequent
    runs give no new data.</p>
  <h2 class="font-semibold">Google Sheets (optional)</h2>
  <p class="text-xs text-slate-500">
    {% if service_account_present %}
    service_account.json found. Give the service-account email editor
    access to your sheet, then save its URL here.
    {% else %}
    To turn this on, put a service_account.json file in the app folder.
    The app runs fully without it.
    {% endif %}</p>
  <input name="sheet_url" value="{{ sheet_url }}"
    placeholder="https://docs.google.com/spreadsheets/d/…"
    class="border rounded px-2 py-1.5">
  <button class="bg-sky-600 hover:bg-sky-500 text-white px-4 py-2
    rounded self-start">Save settings</button>
</form>

<div class="flex flex-col gap-4">
  <div class="bg-white rounded-lg shadow p-5 text-sm">
    <h2 class="font-semibold mb-2">Maintenance</h2>
    <div class="flex flex-wrap gap-2">
      <button onclick="doPost('/rescore')" class="bg-indigo-600
        hover:bg-indigo-500 text-white text-xs px-3 py-1.5 rounded">
        Re-score all leads</button>
      <button onclick="doPost('/verify/start')" class="bg-indigo-700
        hover:bg-indigo-600 text-white text-xs px-3 py-1.5 rounded">
        Verification sweep (all leads)</button>
      <button onclick="doPost('/dedup/scan')" class="bg-sky-700
        hover:bg-sky-600 text-white text-xs px-3 py-1.5 rounded">
        Scan for duplicates</button>
    </div>
    <p id="maintMsg" class="text-xs text-slate-500 mt-2"></p>
  </div>
  <div class="bg-white rounded-lg shadow p-5 text-sm">
    <h2 class="font-semibold mb-2">Publish to Google Sheets</h2>
    <p class="text-xs text-slate-500 mb-2">One-way push of the scored
      lead list to the "Leads" worksheet. The app never reads the sheet
      back.
      {% if last_publish %} Last publish: {{ last_publish }}.{% endif %}</p>
    <button onclick="doPost('/publish')"
      {% if not sheets_ready %}disabled{% endif %}
      class="bg-emerald-600 hover:bg-emerald-500 disabled:bg-slate-300
      text-white text-xs px-3 py-1.5 rounded">Publish now</button>
    {% if not sheets_ready %}
    <p class="text-xs text-slate-400 mt-1">Not available: add
      service_account.json and a Sheet URL first.</p>
    {% endif %}
  </div>
</div>
</div>
{% endblock %}
{% block scripts %}
<script>
function doPost(url) {
  document.getElementById('maintMsg').textContent = 'Started…';
  postAction(url).then(function (r) {
    document.getElementById('maintMsg').textContent =
      r.message || (r.ok ? 'Done.' : 'Failed.');
  });
}
setInterval(function () {
  fetch('/status').then(function (r) { return r.json(); })
  .then(function (s) {
    var parts = [];
    ['verify', 'publish'].forEach(function (k) {
      var job = s.jobs[k];
      if (job && job.running) parts.push(job.label + ': ' + job.message);
    });
    if (parts.length)
      document.getElementById('maintMsg').textContent = parts.join(' | ');
  }).catch(function () {});
}, 2000);
</script>
{% endblock %}"""

LENDERS_HTML = """{% extends "base.html" %}
{% block body %}
<div class="flex items-center justify-between mb-4">
  <h1 class="text-xl font-bold">Lender map
    <span class="text-sm font-normal text-slate-500">
      ({{ lenders|length }} lenders from SBA loan data)</span></h1>
  <a href="{{ url_for('export_lenders') }}" class="bg-emerald-600
    hover:bg-emerald-500 text-white text-sm px-4 py-1.5 rounded">
    Export CSV</a>
</div>
<div class="bg-white rounded-lg shadow overflow-x-auto">
<table class="data w-full text-sm">
<thead class="bg-slate-50 text-left text-xs uppercase text-slate-500">
<tr><th>Lender</th><th>City</th><th>State</th>
<th class="text-right">Loans</th>
<th class="text-right">Total approved</th></tr></thead>
<tbody class="divide-y divide-slate-100">
{% for l in lenders %}
<tr><td class="font-medium">{{ l['name'] }}</td><td>{{ l['city'] }}</td>
<td>{{ l['state'] }}</td>
<td class="text-right">{{ l['loan_count'] }}</td>
<td class="text-right">${{ '{:,}'.format(l['total_dollars']) }}</td></tr>
{% else %}
<tr><td colspan="5" class="text-center text-slate-500 py-8">
  No lender data yet. Run the SBA borrowers connector.</td></tr>
{% endfor %}
</tbody></table></div>
{% endblock %}"""

app.jinja_loader = DictLoader({"base.html": BASE_HTML})


@app.context_processor
def inject_common():
    conn = get_db()
    try:
        dedup_count = conn.execute(
            "SELECT COUNT(*) AS n FROM dedup_queue WHERE resolved = 0"
        ).fetchone()["n"]
    finally:
        conn.close()
    return {
        "dedup_count": dedup_count,
        "type_labels": LEAD_TYPES,
        "compliance_note": COMPLIANCE_NOTE,
    }


# ---------------------------------------------------------------------
# Query helpers
# ---------------------------------------------------------------------

SORTABLE = {
    "firm_name", "lead_type", "city", "state", "score", "tier",
    "contact_count", "last_seen", "phone",
}


def build_lead_filters(args):
    """Turn the request arguments into a WHERE clause."""
    where = []
    params = []
    if args.get("lead_type"):
        where.append("lead_type = ?")
        params.append(args["lead_type"])
    if args.get("tier"):
        where.append("tier = ?")
        params.append(args["tier"])
    if args.get("state"):
        where.append("upper(state) = ?")
        params.append(args["state"].upper())
    if args.get("q"):
        like = f"%{args['q'].strip()}%"
        where.append(
            "(firm_name LIKE ? OR dba_name LIKE ? OR contact_name LIKE ?"
            " OR city LIKE ? OR phone LIKE ? OR email LIKE ?"
            " OR website LIKE ?)"
        )
        params.extend([like] * 7)
    clause = (" WHERE " + " AND ".join(where)) if where else ""
    return clause, params


def filtered_lead_ids(conn, args):
    clause, params = build_lead_filters(args)
    return [
        r["id"]
        for r in conn.execute("SELECT id FROM all_leads" + clause, params)
    ]


# ---------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------

@app.route("/")
def page_dashboard():
    conn = get_db()
    try:
        total = conn.execute(
            "SELECT COUNT(*) AS n FROM leads"
        ).fetchone()["n"]
        tiers = {
            r["tier"]: r["n"]
            for r in conn.execute(
                "SELECT tier, COUNT(*) AS n FROM leads GROUP BY tier"
            )
        }
        by_type = conn.execute(
            "SELECT lead_type, COUNT(*) AS n FROM leads "
            "GROUP BY lead_type ORDER BY n DESC"
        ).fetchall()
        contact_total = conn.execute(
            "SELECT COUNT(*) AS n FROM contacts"
        ).fetchone()["n"]
        last_runs = {}
        for c in AUTO_CONNECTORS:
            row = conn.execute(
                "SELECT * FROM runs WHERE connector_id = ? "
                "ORDER BY id DESC LIMIT 1",
                (c["id"],),
            ).fetchone()
            if row:
                last_runs[c["id"]] = row
        target_state = get_target_state(conn)
    finally:
        conn.close()
    return render_template_string(
        DASHBOARD_HTML,
        total=total,
        tiers=tiers,
        by_type=by_type,
        contact_total=contact_total,
        connectors=CONNECTORS,
        last_runs=last_runs,
        target_state=target_state,
    )


@app.route("/leads")
def page_leads():
    conn = get_db()
    try:
        clause, params = build_lead_filters(request.args)
        result_count = conn.execute(
            "SELECT COUNT(*) AS n FROM all_leads" + clause, params
        ).fetchone()["n"]
        sort = request.args.get("sort", "score")
        if sort not in SORTABLE:
            sort = "score"
        direction = (
            "ASC" if request.args.get("dir") == "asc" else "DESC"
        )
        leads = conn.execute(
            f"SELECT * FROM all_leads{clause} "
            f"ORDER BY {sort} {direction}, firm_name ASC LIMIT 200",
            params,
        ).fetchall()
    finally:
        conn.close()

    def sort_url(col):
        args = dict(request.args)
        if args.get("sort", "score") == col:
            args["dir"] = "asc" if args.get("dir") != "asc" else "desc"
        else:
            args["dir"] = "desc"
        args["sort"] = col
        return url_for("page_leads", **args)

    return render_template_string(
        LEADS_HTML,
        leads=leads,
        result_count=result_count,
        sort_url=sort_url,
    )


@app.route("/dedup")
def page_dedup():
    conn = get_db()
    try:
        pairs = []
        for row in conn.execute(
            "SELECT * FROM dedup_queue WHERE resolved = 0 ORDER BY id"
        ):
            a = conn.execute(
                "SELECT * FROM leads WHERE id = ?", (row["lead_id_a"],)
            ).fetchone()
            b = conn.execute(
                "SELECT * FROM leads WHERE id = ?", (row["lead_id_b"],)
            ).fetchone()
            if a is None or b is None:
                conn.execute(
                    "UPDATE dedup_queue SET resolved = 1, "
                    "resolution = 'gone' WHERE id = ?",
                    (row["id"],),
                )
                continue
            keep = a if count_non_empty(a) >= count_non_empty(b) else b
            pairs.append({
                "id": row["id"],
                "reason": row["match_reason"],
                "similarity": row["similarity"],
                "a": a,
                "b": b,
                "keep_id": keep["id"],
            })
        conn.commit()
    finally:
        conn.close()
    return render_template_string(DEDUP_HTML, pairs=pairs)


@app.route("/import")
def page_import():
    return render_template_string(
        IMPORT_HTML, preselect=request.args.get("type", "generic")
    )


IMPORT_GUESSES = {
    "firm_name": ["firm", "company", "business", "organization", "name"],
    "dba_name": ["dba"],
    "contact_name": ["contact", "person", "agent name", "full name"],
    "address": ["address", "street"],
    "city": ["city"],
    "state": ["state"],
    "zip_code": ["zip", "postal"],
    "phone": ["phone", "tel"],
    "email": ["email", "e-mail"],
    "website": ["website", "url", "web"],
    "external_id": ["license", "id", "crd", "number"],
    "aum_dollars": ["aum", "assets"],
    "num_employees": ["employee", "staff"],
    "source_detail": ["source", "detail"],
    "notes": ["note", "comment"],
}


@app.route("/import/upload", methods=["POST"])
def import_upload():
    file = request.files.get("file")
    lead_type = request.form.get("lead_type", "generic")
    if lead_type not in LEAD_TYPES:
        lead_type = "generic"
    if file is None or not file.filename:
        return redirect(url_for("page_import", msg="Pick a CSV file first."))
    IMPORT_DIR.mkdir(exist_ok=True)
    token = hashlib.sha1(
        f"{file.filename}{time.time()}".encode()
    ).hexdigest()[:16]
    saved = IMPORT_DIR / f"{token}.csv"
    file.save(saved)
    try:
        columns, sample, row_count = _read_csv_head(saved)
    except Exception as exc:
        saved.unlink(missing_ok=True)
        return redirect(url_for(
            "page_import",
            msg=f"Could not read the file as CSV: {exc}",
        ))
    (IMPORT_DIR / f"{token}.meta").write_text(
        json.dumps({"filename": file.filename, "lead_type": lead_type}),
        encoding="utf-8",
    )
    guesses = {}
    lowered = [c.lower() for c in columns]
    for field, cands in IMPORT_GUESSES.items():
        for cand in cands:
            hit = next(
                (i for i, col in enumerate(lowered) if cand in col), None
            )
            if hit is not None and hit not in guesses.values():
                guesses[field] = hit
                break
    return render_template_string(
        MAPPING_HTML,
        token=token,
        filename=file.filename,
        lead_type=lead_type,
        columns=columns,
        sample=sample,
        row_count=row_count,
        guesses=guesses,
        import_fields=IMPORT_FIELDS,
        field_list=IMPORT_FIELDS,
    )


def _read_csv_head(path, sample_size=10):
    """Read the header, a sample, and the row count of a CSV file."""
    with open(path, encoding="utf-8-sig", errors="replace", newline="") as fh:
        first = fh.readline()
        if not first.strip():
            raise ValueError("the first row is empty")
        delim = sniff_delimiter(first)
        columns = next(csv.reader([first], delimiter=delim))
        sample = []
        count = 0
        for row in csv.reader(fh, delimiter=delim):
            if not any(cell.strip() for cell in row):
                continue
            count += 1
            if len(sample) < sample_size:
                sample.append(row)
    return columns, sample, count


@app.route("/import/commit", methods=["POST"])
def import_commit():
    token = re.sub(r"[^0-9a-f]", "", request.form.get("token", ""))
    saved = IMPORT_DIR / f"{token}.csv"
    meta_path = IMPORT_DIR / f"{token}.meta"
    if not token or not saved.exists():
        return redirect(url_for(
            "page_import", msg="The upload expired. Upload the file again."
        ))
    meta = json.loads(meta_path.read_text(encoding="utf-8"))
    lead_type = request.form.get("lead_type", meta.get("lead_type", "generic"))
    if lead_type not in LEAD_TYPES:
        lead_type = "generic"
    mapping = {}
    for field, _label in IMPORT_FIELDS:
        raw = request.form.get(f"map_{field}", "")
        if raw.isdigit():
            mapping[field] = int(raw)
    if "firm_name" not in mapping and "contact_name" not in mapping:
        return redirect(url_for(
            "page_import",
            msg="Map at least the firm name or the contact name.",
        ))
    stats = {"seen": 0, "added": 0, "updated": 0, "deduped": 0}
    conn = get_db()
    try:
        with open(
            saved, encoding="utf-8-sig", errors="replace", newline=""
        ) as fh:
            first = fh.readline()
            delim = sniff_delimiter(first)
            for row in csv.reader(fh, delimiter=delim):
                if not any(cell.strip() for cell in row):
                    continue
                lead = {
                    "lead_type": lead_type,
                    "source_connector": "csv_import",
                    "source_detail": f"CSV import: {meta.get('filename')}",
                }
                for field, idx in mapping.items():
                    value = row[idx].strip() if idx < len(row) else ""
                    if field in ("aum_dollars", "num_employees"):
                        lead[field] = to_int(value)
                    else:
                        lead[field] = value
                if not lead.get("firm_name"):
                    lead["firm_name"] = lead.get("contact_name", "")
                if not lead.get("firm_name"):
                    continue
                stats["seen"] += 1
                _, action = upsert_lead(conn, lead)
                stats[action] += 1
        conn.commit()
        run_fuzzy_dedup(conn)
    finally:
        conn.close()
        saved.unlink(missing_ok=True)
        meta_path.unlink(missing_ok=True)
    return redirect(url_for(
        "page_leads",
        lead_type=lead_type,
        msg=(
            f"Import done: {stats['seen']} rows, {stats['added']} new, "
            f"{stats['updated'] + stats['deduped']} matched existing "
            "leads."
        ),
    ))


@app.route("/settings", methods=["GET", "POST"])
def page_settings():
    conn = get_db()
    try:
        if request.method == "POST":
            old_weights = get_weights(conn)
            for key in DEFAULT_WEIGHTS:
                raw = request.form.get(key, "")
                if raw != "":
                    set_setting(conn, key, to_int(raw) or 0)
            state = request.form.get("target_state", "WI")
            if state in SUPPORTED_STATES:
                set_setting(conn, "target_state", state)
            set_setting(
                conn, "schedule_enabled",
                "1" if request.form.get("schedule_enabled") else "0",
            )
            day = request.form.get("schedule_day", "mon")
            if day in ("mon", "tue", "wed", "thu", "fri", "sat", "sun"):
                set_setting(conn, "schedule_day", day)
            set_setting(
                conn, "sheet_url", request.form.get("sheet_url", "").strip()
            )
            conn.commit()
            changed = old_weights != get_weights(conn)
            if changed:
                rescore_all(conn)
            apply_schedule()
            return redirect(url_for(
                "page_settings",
                msg="Settings saved."
                    + (" All scores were computed again." if changed else ""),
            ))
        weights = get_weights(conn)
        context = {
            "weights": weights,
            "weight_labels": WEIGHT_LABELS,
            "target_state": get_target_state(conn),
            "supported_states": SUPPORTED_STATES,
            "state_names": STATE_NAMES,
            "schedule_enabled": get_setting(conn, "schedule_enabled") == "1",
            "schedule_day": get_setting(conn, "schedule_day", "mon"),
            "sheet_url": get_setting(conn, "sheet_url"),
            "last_publish": get_setting(conn, "last_publish"),
            "service_account_present": SERVICE_ACCOUNT_PATH.exists(),
            "sheets_ready": sheets_available(conn),
        }
    finally:
        conn.close()
    return render_template_string(SETTINGS_HTML, **context)


@app.route("/lenders")
def page_lenders():
    conn = get_db()
    try:
        lenders = conn.execute(
            "SELECT * FROM lenders ORDER BY loan_count DESC, name"
        ).fetchall()
    finally:
        conn.close()
    return render_template_string(LENDERS_HTML, lenders=lenders)


# ---------------------------------------------------------------------
# JSON API
# ---------------------------------------------------------------------

@app.route("/status")
def api_status():
    conn = get_db()
    try:
        dedup_pending = conn.execute(
            "SELECT COUNT(*) AS n FROM dedup_queue WHERE resolved = 0"
        ).fetchone()["n"]
        total = conn.execute(
            "SELECT COUNT(*) AS n FROM leads"
        ).fetchone()["n"]
    finally:
        conn.close()
    return jsonify({
        "jobs": job_snapshot(),
        "dedup_pending": dedup_pending,
        "total_leads": total,
    })


@app.route("/run/<connector_id>", methods=["POST"])
def api_run_connector(connector_id):
    ok, message = start_connector_run(connector_id)
    return jsonify({"ok": ok, "message": message})


@app.route("/verify/start", methods=["POST"])
def api_verify_start():
    conn = get_db()
    try:
        ids = filtered_lead_ids(conn, request.args)
    finally:
        conn.close()
    if not ids:
        return jsonify({"ok": False, "message": "No leads match."})
    ok, message = start_verify_sweep(ids)
    return jsonify({"ok": ok, "message": message})


@app.route("/scrape/bulk", methods=["POST"])
def api_scrape_bulk():
    body = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        ids = filtered_lead_ids(conn, request.args)
    finally:
        conn.close()
    if not ids:
        return jsonify({"ok": False, "message": "No leads match."})
    ok, message = start_contact_scrape(
        ids, only_unscraped=bool(body.get("only_unscraped"))
    )
    return jsonify({"ok": ok, "message": message})


@app.route("/rescore", methods=["POST"])
def api_rescore():
    conn = get_db()
    try:
        n = rescore_all(conn)
    finally:
        conn.close()
    return jsonify({"ok": True, "message": f"{n} leads scored again."})


@app.route("/dedup/scan", methods=["POST"])
def api_dedup_scan():
    conn = get_db()
    try:
        added = run_fuzzy_dedup(conn)
    finally:
        conn.close()
    return jsonify({
        "ok": True,
        "message": f"{added} new candidate pairs found.",
    })


@app.route("/dedup/<int:pair_id>/merge", methods=["POST"])
def api_dedup_merge(pair_id):
    conn = get_db()
    try:
        pair = conn.execute(
            "SELECT * FROM dedup_queue WHERE id = ? AND resolved = 0",
            (pair_id,),
        ).fetchone()
        if pair is None:
            return jsonify({"ok": False, "message": "Pair not found."})
        a = conn.execute(
            "SELECT * FROM leads WHERE id = ?", (pair["lead_id_a"],)
        ).fetchone()
        b = conn.execute(
            "SELECT * FROM leads WHERE id = ?", (pair["lead_id_b"],)
        ).fetchone()
        if a is None or b is None:
            conn.execute(
                "UPDATE dedup_queue SET resolved = 1, resolution = 'gone' "
                "WHERE id = ?",
                (pair_id,),
            )
            conn.commit()
            return jsonify({"ok": False, "message": "A lead is gone."})
        if count_non_empty(a) >= count_non_empty(b):
            keep, drop = a, b
        else:
            keep, drop = b, a
        merge_leads(conn, keep["id"], drop["id"])
        compute_score(conn, keep["id"])
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "message": "Merged."})


@app.route("/dedup/<int:pair_id>/distinct", methods=["POST"])
def api_dedup_distinct(pair_id):
    conn = get_db()
    try:
        conn.execute(
            "UPDATE dedup_queue SET resolved = 1, resolution = 'distinct' "
            "WHERE id = ?",
            (pair_id,),
        )
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "message": "Kept as two leads."})


@app.route("/api/lead/<int:lead_id>")
def api_lead_detail(lead_id):
    conn = get_db()
    try:
        lead = conn.execute(
            "SELECT * FROM leads WHERE id = ?", (lead_id,)
        ).fetchone()
        if lead is None:
            return jsonify({"error": "not found"}), 404
        contacts = [
            dict(r)
            for r in conn.execute(
                "SELECT * FROM contacts WHERE lead_id = ? "
                "ORDER BY confidence DESC, name",
                (lead_id,),
            )
        ]
        checks = {
            r["check_name"]: r
            for r in conn.execute(
                "SELECT * FROM verifications WHERE lead_id = ?", (lead_id,)
            )
        }
        weights = get_weights(conn)
        breakdown = []
        for check_name, weight_key in CHECK_TO_WEIGHT.items():
            row = checks.get(check_name)
            breakdown.append({
                "label": WEIGHT_LABELS[weight_key],
                "weight": weights[weight_key],
                "passed": bool(row and row["passed"]),
                "detail": row["detail"] if row else "Not checked yet.",
                "checked_at": row["checked_at"] if row else "",
            })
    finally:
        conn.close()
    return jsonify({
        "lead": dict(lead),
        "type_label": LEAD_TYPES.get(lead["lead_type"], lead["lead_type"]),
        "contacts": contacts,
        "breakdown": breakdown,
    })


@app.route("/api/lead/<int:lead_id>/verify", methods=["POST"])
def api_lead_verify(lead_id):
    conn = get_db()
    try:
        lead = conn.execute(
            "SELECT * FROM leads WHERE id = ?", (lead_id,)
        ).fetchone()
        if lead is None:
            return jsonify({"ok": False, "message": "Lead not found."})
        verify_lead(conn, lead, online=True)
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "message": "Checks done."})


@app.route("/api/lead/<int:lead_id>/scrape", methods=["POST"])
def api_lead_scrape(lead_id):
    conn = get_db()
    try:
        lead = conn.execute(
            "SELECT * FROM leads WHERE id = ?", (lead_id,)
        ).fetchone()
    finally:
        conn.close()
    if lead is None:
        return jsonify({"ok": False, "message": "Lead not found."})
    if not lead["website"]:
        return jsonify({
            "ok": False,
            "message": "This lead has no website on file.",
        })
    ok, message = start_contact_scrape([lead_id])
    return jsonify({"ok": ok, "message": message})


@app.route("/publish", methods=["POST"])
def api_publish():
    ok, message = start_publish()
    return jsonify({"ok": ok, "message": message})


@app.route("/quit", methods=["POST"])
def api_quit():
    threading.Timer(0.4, os._exit, args=(0,)).start()
    return jsonify({"ok": True, "message": "The server stops now."})


# ---------------------------------------------------------------------
# CSV exports
# ---------------------------------------------------------------------

def _csv_response(rows, header, kind):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(rows)
    stamp = datetime.now().strftime("%Y%m%d")
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition":
                f"attachment; filename={kind}_{stamp}.csv"
        },
    )


EXPORT_LEAD_COLUMNS = [
    "id", "lead_type", "external_id", "firm_name", "dba_name",
    "contact_name", "address", "city", "state", "zip_code", "phone",
    "email", "website", "aum_dollars", "num_employees",
    "source_connector", "source_detail", "corroborators", "score",
    "tier", "status", "first_seen", "last_seen", "contact_count",
    "notes",
]


@app.route("/export/leads.csv")
def export_leads():
    conn = get_db()
    try:
        clause, params = build_lead_filters(request.args)
        rows = conn.execute(
            "SELECT * FROM all_leads" + clause +
            " ORDER BY score DESC, firm_name",
            params,
        ).fetchall()
    finally:
        conn.close()
    data = [
        [r[c] for c in EXPORT_LEAD_COLUMNS] for r in rows
    ]
    return _csv_response(data, EXPORT_LEAD_COLUMNS, "leads")


@app.route("/export/contacts.csv")
def export_contacts():
    conn = get_db()
    try:
        clause, params = build_lead_filters(request.args)
        rows = conn.execute(
            """SELECT c.*, l.firm_name, l.lead_type, l.city AS lead_city,
                l.state AS lead_state, l.score, l.tier
               FROM contacts c
               JOIN (SELECT * FROM all_leads""" + clause + """) l
                 ON l.id = c.lead_id
               ORDER BY l.score DESC, l.firm_name, c.confidence DESC""",
            params,
        ).fetchall()
    finally:
        conn.close()
    header = [
        "firm_name", "lead_type", "lead_city", "lead_state", "score",
        "tier", "name", "title", "email", "phone", "linkedin_url",
        "source_url", "confidence", "scraped_at",
    ]
    data = [[r[c] for c in header] for r in rows]
    return _csv_response(data, header, "contacts")


@app.route("/export/lenders.csv")
def export_lenders():
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT name, city, state, loan_count, total_dollars, "
            "last_seen FROM lenders ORDER BY loan_count DESC"
        ).fetchall()
    finally:
        conn.close()
    header = ["name", "city", "state", "loan_count", "total_dollars",
              "last_seen"]
    data = [[r[c] for c in header] for r in rows]
    return _csv_response(data, header, "lenders")


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Lead Engine v2")
    parser.add_argument(
        "--demo", action="store_true",
        help="Put demo data in an empty database.",
    )
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    args = parser.parse_args()

    init_db()
    migrate_predecessor()
    if args.demo:
        seed_demo()
    try:
        apply_schedule()
    except Exception as exc:
        print(f"Scheduler not started: {exc}", file=sys.stderr)

    app.run(host=HOST, port=args.port, threaded=True, use_reloader=False)


if __name__ == "__main__":
    main()
